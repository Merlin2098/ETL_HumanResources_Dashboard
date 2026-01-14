# ui/etls/bd/worker.py
"""
Worker para ETL de Base de Datos
Ejecuta: Bronze → Silver → Centros de Costo → Gold → Flags

ARQUITECTURA:
- Step 1: Bronze → Silver (Excel con headers en fila 10)
- Step 1.5: Extracción de Centros de Costo (versionado dual: /actual + /historico)
- Step 2: Silver → Gold (Empleados + Practicantes)
- Step 3: Aplicación de Flags (opcional, requiere archivos CC)

Última modificación: 14.01.2025 - Corrección patrón dual /actual + /historico en step3
"""
from pathlib import Path
from typing import Dict
import sys
import time
import json
import re
from datetime import datetime

# Asegurar que el directorio raíz del proyecto esté en el path
project_root = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(project_root))

from ui.workers.base_worker import BaseETLWorker


class BDWorker(BaseETLWorker):
    """Worker para procesamiento de BD con lazy loading"""
    
    def __init__(self, archivos, output_dir):
        """
        Args:
            archivos: List[Path] - Lista con 1 archivo Excel
            output_dir: Path - Directorio de salida (carpeta de trabajo)
        """
        super().__init__(archivos, output_dir)
        
        # Lazy loading de DuckDB (solo si se ejecuta step 3)
        self.duckdb = None
        
        # Timers
        self.timers = {
            'total': 0,
            'step1': 0,
            'step1.5': 0,
            'step2': 0,
            'step3': 0
        }
    
    def get_pipeline_name(self) -> str:
        return "bd"
    
    def execute_etl(self) -> Dict:
        """
        Ejecuta el ETL completo de BD (Steps 1 + 1.5 + 2 + 3)
        
        Returns:
            dict con resultados del proceso
        """
        tiempo_inicio_total = time.time()
        
        try:
            return self._ejecutar_etl_completo(tiempo_inicio_total)
                
        except Exception as e:
            self.logger.error(f"❌ Error crítico en ETL: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
            
            self.timers['total'] = time.time() - tiempo_inicio_total
            
            return {
                'success': False,
                'error': str(e),
                'timers': self.timers
            }
    
    # ========================================================================
    # MODO: ETL COMPLETO (Steps 1 + 1.5 + 2)
    # ========================================================================
    
    def _ejecutar_etl_completo(self, tiempo_inicio_total) -> Dict:
        """Ejecuta Steps 1 + 1.5 + 2 + 3"""
        resultado = {}
        
        # Validar que haya archivo
        if not self.archivos or len(self.archivos) == 0:
            raise ValueError("No se proporcionó archivo Excel")
        
        archivo_excel = self.archivos[0]  # BD usa solo 1 archivo
        
        self.logger.info("="*70)
        self.logger.info("PROCESAMIENTO BD - ETL COMPLETO")
        self.logger.info("="*70)
        self.logger.info(f"📂 Archivo: {archivo_excel.name}")
        self.logger.info(f"📁 Carpeta de trabajo: {self.output_dir}")
        self.logger.info("")
        
        # Step 1: Bronze → Silver
        self.progress_updated.emit(5, "🔄 Step 1: Bronze → Silver")
        resultado['step1'] = self._step1_bronze_to_silver(archivo_excel)
        self.progress_updated.emit(25, "✓ Step 1 completado")
        
        # Step 1.5: Extracción de Centros de Costo
        self.progress_updated.emit(30, "🔄 Step 1.5: Centros de Costo")
        resultado['step1.5'] = self._step1_5_extraer_centros_costo(resultado['step1']['parquet'])
        self.progress_updated.emit(50, "✓ Step 1.5 completado")
        
        # Step 2: Silver → Gold
        self.progress_updated.emit(55, "🔄 Step 2: Silver → Gold")
        resultado['step2'] = self._step2_silver_to_gold(resultado['step1']['parquet'])
        self.progress_updated.emit(75, "✓ Step 2 completado")
        
        # Step 3: Aplicación de Flags (usa el CC_ACTUAL recién generado)
        self.progress_updated.emit(80, "🔄 Step 3: Aplicando Flags")
        try:
            resultado['step3'] = self._step3_aplicar_flags_automatico(resultado['step1.5']['parquet_actual'])
            self.progress_updated.emit(100, "✓ Step 3 completado")
        except Exception as e:
            self.logger.warning(f"⚠️  Step 3 (Flags) falló: {str(e)}")
            resultado['step3'] = {'error': str(e)}
            self.progress_updated.emit(100, "⚠️  ETL completado (Step 3 con errores)")
        
        # Resultado final
        self.timers['total'] = time.time() - tiempo_inicio_total
        resultado['success'] = True
        resultado['timers'] = self.timers
        
        mensaje_base = (
            f"ETL Completo finalizado:\n"
            f"  • Silver: {resultado['step1']['registros']:,} registros\n"
            f"  • Centros de Costo: {resultado['step1.5']['registros']:,} únicos\n"
            f"  • Empleados: {resultado['step2']['empleados']:,} registros\n"
            f"  • Practicantes: {resultado['step2']['practicantes']:,} registros\n"
        )
        
        if 'error' not in resultado['step3']:
            mensaje_base += f"  • Flags aplicados: {resultado['step3']['registros']:,} registros\n"
        else:
            mensaje_base += f"  ⚠️  Flags: Error (ver log)\n"
        
        mensaje_base += f"  ⏱️  Tiempo total: {self.logger.format_duration(self.timers['total'])}"
        
        resultado['mensaje'] = mensaje_base
        self.logger.info("\n" + "="*70)
        self.logger.info(mensaje_base)
        self.logger.info("="*70)
        
        return resultado
    
    # ========================================================================
    # STEP 1: BRONZE → SILVER
    # ========================================================================
    
    def _step1_bronze_to_silver(self, archivo_excel):
        """Procesa Excel Bronze → Parquet Silver"""
        tiempo_inicio = time.time()
        
        self.logger.info("="*70)
        self.logger.info("STEP 1: BRONZE → SILVER")
        self.logger.info("="*70)
        self.logger.info(f"📂 Archivo: {archivo_excel.name}")
        
        # Lazy load de openpyxl y polars
        import openpyxl
        import polars as pl
        
        # Extraer datos
        self.logger.info("📊 Extrayendo datos del Excel...")
        headers, data_rows = self._extraer_datos_excel(archivo_excel, openpyxl)
        
        # Crear DataFrame
        self.logger.info("🔄 Creando DataFrame...")
        df = self._crear_dataframe_polars(headers, data_rows, pl)
        
        # Guardar Silver
        self.logger.info("💾 Guardando en capa Silver...")
        ruta_parquet, ruta_excel = self._guardar_silver(df, self.output_dir)
        
        self.timers['step1'] = time.time() - tiempo_inicio
        
        self.logger.info("-"*70)
        self.logger.info(f"✓ Step 1 completado")
        self.logger.info(f"  • Registros: {len(df):,}")
        self.logger.info(f"  • Columnas: {len(df.columns)}")
        self.logger.info(f"  • Parquet: {ruta_parquet.name}")
        self.logger.info(f"  ⏱️  Duración: {self.logger.format_duration(self.timers['step1'])}")
        self.logger.info("-"*70)
        
        return {
            'registros': len(df),
            'columnas': len(df.columns),
            'parquet': ruta_parquet,
            'excel': ruta_excel
        }
    
    def _extraer_datos_excel(self, archivo_excel, openpyxl):
        """Extrae headers y datos del Excel (fila 10 header, datos desde 11)"""
        wb = openpyxl.load_workbook(archivo_excel, read_only=True, data_only=True)
        ws = wb.active
        
        HEADER_ROW = 10
        DATA_START_ROW = 11
        
        # Leer headers (fila 10)
        headers = [cell.value for cell in ws[HEADER_ROW] if cell.value is not None]
        self.logger.info(f"  ✓ Headers encontrados: {len(headers)} columnas")
        
        # Leer datos (desde fila 11)
        data_rows = []
        for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
            if any(cell is not None for cell in row[:len(headers)]):
                data_rows.append(row[:len(headers)])
        
        wb.close()
        
        self.logger.info(f"  ✓ Filas de datos: {len(data_rows):,}")
        
        return headers, data_rows
    
    def _crear_dataframe_polars(self, headers, data_rows, pl):
        """Crea DataFrame Polars desde headers y datos"""
        # Crear diccionario para cada columna
        data_dict = {header: [] for header in headers}
        
        for row in data_rows:
            for idx, value in enumerate(row):
                header = headers[idx]
                # Convertir todo a String para evitar conflictos de tipo
                data_dict[header].append(str(value) if value is not None else None)
        
        # Crear DataFrame
        df = pl.DataFrame(data_dict)
        
        self.logger.info(f"  ✓ DataFrame creado: {df.shape[0]:,} filas × {df.shape[1]} columnas")
        
        return df
    
    def _guardar_silver(self, df, carpeta_trabajo):
        """Guarda DataFrame en carpeta silver/"""
        carpeta_silver = carpeta_trabajo / "silver"
        carpeta_silver.mkdir(exist_ok=True)
        
        # Parquet (sin timestamp)
        ruta_parquet = carpeta_silver / "bd_silver.parquet"
        df.write_parquet(ruta_parquet, compression="snappy")
        self.logger.info(f"  ✓ Parquet: {ruta_parquet.name}")
        
        # Excel (opcional)
        ruta_excel = carpeta_silver / "bd_silver.xlsx"
        df.write_excel(ruta_excel)
        self.logger.info(f"  ✓ Excel: {ruta_excel.name}")
        
        return ruta_parquet, ruta_excel
    
    # ========================================================================
    # STEP 1.5: EXTRACCIÓN CENTROS DE COSTO
    # ========================================================================
    
    def _step1_5_extraer_centros_costo(self, ruta_silver):
        """
        Extrae y procesa Centros de Costo directamente desde Silver usando Polars.
        Replica la lógica de step1.5_centrosdecosto.py
        """
        tiempo_inicio = time.time()
        
        self.logger.info("")
        self.logger.info("="*70)
        self.logger.info("STEP 1.5: EXTRACCIÓN DE CENTROS DE COSTO")
        self.logger.info("="*70)
        
        import polars as pl
        
        # Cargar datos Silver
        self.logger.info("📊 Cargando datos Silver...")
        df_silver = pl.read_parquet(ruta_silver)
        self.logger.info(f"  ✓ Silver: {len(df_silver):,} registros")
        
        # Cargar esquema de CC
        self.logger.info("📋 Cargando esquema de Centros de Costo...")
        ruta_esquema = self._buscar_esquema('esquema_cc.json')
        
        if not ruta_esquema:
            raise FileNotFoundError("No se encontró esquema_cc.json")
        
        import json
        with open(ruta_esquema, 'r', encoding='utf-8') as f:
            esquema = json.load(f)
        
        columnas_requeridas = esquema['columnas_requeridas']
        columna_dedupe = esquema['columna_deduplicacion']
        
        self.logger.info(f"  ✓ Columnas a extraer: {len(columnas_requeridas)}")
        
        # Verificar que existan las columnas
        columnas_faltantes = set(columnas_requeridas) - set(df_silver.columns)
        if columnas_faltantes:
            raise ValueError(f"Columnas faltantes: {columnas_faltantes}")
        
        # Extraer y deduplicar
        self.logger.info("🔄 Procesando centros de costo...")
        df_cc = df_silver.select(columnas_requeridas)
        
        registros_antes = len(df_cc)
        df_cc = df_cc.unique(subset=[columna_dedupe], keep='first')
        df_cc = df_cc.sort(columna_dedupe)
        
        self.logger.info(f"  ✓ Antes de deduplicación: {registros_antes:,}")
        self.logger.info(f"  ✓ Después de deduplicación: {len(df_cc):,}")
        self.logger.info(f"  ✓ Centros de costo únicos: {df_cc[columna_dedupe].n_unique()}")
        
        # Guardar resultado con versionamiento en carpeta SEPARADA
        carpeta_trabajo = ruta_silver.parent.parent
        carpeta_cc = carpeta_trabajo / "centros_costo"
        carpeta_actual = carpeta_cc / "actual"
        carpeta_historico = carpeta_cc / "historico"
        
        carpeta_actual.mkdir(parents=True, exist_ok=True)
        carpeta_historico.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Actual (sin timestamp)
        nombre_base_actual = "CC_ACTUAL"
        ruta_actual = carpeta_actual / f"{nombre_base_actual}.parquet"
        df_cc.write_parquet(ruta_actual, compression="snappy")
        self.logger.info(f"  ✓ CC (actual): {ruta_actual.name}")
        
        ruta_excel = carpeta_actual / f"{nombre_base_actual}.xlsx"
        df_cc.write_excel(ruta_excel)
        
        # Histórico (con timestamp)
        nombre_base_historico = f"CC_ACTUAL_{timestamp}"
        ruta_historico = carpeta_historico / f"{nombre_base_historico}.parquet"
        df_cc.write_parquet(ruta_historico, compression="snappy")
        self.logger.info(f"  ✓ CC (histórico): {ruta_historico.name}")
        
        self.timers['step1.5'] = time.time() - tiempo_inicio
        
        self.logger.info("-"*70)
        self.logger.info(f"✓ Step 1.5 completado")
        self.logger.info(f"  • Registros: {len(df_cc):,}")
        self.logger.info(f"  • Archivo actual: {ruta_actual.name}")
        self.logger.info(f"  ⏱️  Duración: {self.logger.format_duration(self.timers['step1.5'])}")
        self.logger.info("-"*70)
        
        return {
            'registros': len(df_cc),
            'parquet_actual': ruta_actual,
            'parquet_historico': ruta_historico,
            'duracion': self.timers['step1.5']
        }
    
    # ========================================================================
    # STEP 2: SILVER → GOLD
    # ========================================================================
    
    def _step2_silver_to_gold(self, ruta_silver):
        """Procesa Silver → Gold (Empleados + Practicantes)"""
        tiempo_inicio = time.time()
        
        self.logger.info("")
        self.logger.info("="*70)
        self.logger.info("STEP 2: SILVER → GOLD")
        self.logger.info("="*70)
        
        import polars as pl
        
        # Cargar Silver
        self.logger.info("📊 Cargando Silver...")
        df_silver = pl.read_parquet(ruta_silver)
        self.logger.info(f"  ✓ Silver: {len(df_silver):,} registros × {len(df_silver.columns)} columnas")
        
        # Cargar esquema
        self.logger.info("📋 Cargando esquema de validación...")
        ruta_esquema = self._buscar_esquema('esquema_bd.json')
        
        if not ruta_esquema:
            raise FileNotFoundError("No se encontró esquema_bd.json")
        
        import json
        with open(ruta_esquema, 'r', encoding='utf-8') as f:
            esquema = json.load(f)
        
        self.logger.info(f"  ✓ Esquema: {esquema['schema_name']} v{esquema['version']}")
        
        # Filtrar columnas
        self.logger.info("🔍 Filtrando columnas según esquema...")
        df_gold = self._filtrar_columnas_gold(df_silver, esquema)
        
        # Convertir fechas
        self.logger.info("📅 Convirtiendo columnas de fecha...")
        df_gold = self._convertir_fechas_gold(df_gold, esquema, pl)
        
        # Dividir por modalidad
        self.logger.info("🔀 Dividiendo por modalidad de contrato...")
        df_empleados, df_practicantes = self._dividir_por_modalidad(df_gold, pl)
        
        # Guardar Gold
        self.logger.info("💾 Guardando en capa Gold...")
        carpeta_trabajo = ruta_silver.parent.parent
        self._guardar_gold(df_empleados, df_practicantes, carpeta_trabajo)
        
        self.timers['step2'] = time.time() - tiempo_inicio
        
        self.logger.info("-"*70)
        self.logger.info(f"✓ Step 2 completado")
        self.logger.info(f"  • Empleados: {len(df_empleados):,}")
        self.logger.info(f"  • Practicantes: {len(df_practicantes):,}")
        self.logger.info(f"  ⏱️  Duración: {self.logger.format_duration(self.timers['step2'])}")
        self.logger.info("-"*70)
        
        return {
            'empleados': len(df_empleados),
            'practicantes': len(df_practicantes),
            'duracion': self.timers['step2']
        }
    
    def _filtrar_columnas_gold(self, df, esquema):
        """Filtra columnas según esquema"""
        required_columns = [col["name"] for col in esquema["columns"]]
        columns_to_select = [col for col in df.columns if col in required_columns]
        
        df_filtered = df.select(columns_to_select)
        
        missing = [col for col in required_columns if col not in df.columns]
        if missing:
            self.logger.warning(f"  ⚠️  Columnas faltantes: {', '.join(missing)}")
        
        self.logger.info(f"  ✓ Columnas filtradas: {len(df_filtered.columns)}")
        
        return df_filtered
    
    def _convertir_fechas_gold(self, df, esquema, pl):
        """Convierte columnas de fecha según esquema"""
        date_columns = [col["name"] for col in esquema["columns"] if col["type"] == "date"]
        
        for col_name in date_columns:
            if col_name in df.columns:
                try:
                    df = df.with_columns(
                        pl.col(col_name)
                        .str.to_datetime(format="%Y-%m-%d %H:%M:%S", strict=False)
                        .cast(pl.Date, strict=False)
                        .alias(col_name)
                    )
                    self.logger.info(f"  ✓ {col_name} convertida a Date")
                except Exception as e:
                    self.logger.warning(f"  ⚠️  Error en {col_name}: {str(e)}")
        
        return df
    
    def _dividir_por_modalidad(self, df, pl):
        """Divide en Empleados y Practicantes según Modalidad de Contrato"""
        if "Modalidad de Contrato" not in df.columns:
            raise ValueError("Columna 'Modalidad de Contrato' no encontrada")
        
        df_practicantes = df.filter(
            pl.col("Modalidad de Contrato").str.contains("TERMINO DE CONVENIO")
        )
        
        df_empleados = df.filter(
            ~pl.col("Modalidad de Contrato").str.contains("TERMINO DE CONVENIO")
        )
        
        self.logger.info(f"  ✓ Practicantes: {df_practicantes.height:,}")
        self.logger.info(f"  ✓ Empleados: {df_empleados.height:,}")
        
        return df_empleados, df_practicantes
    
    def _guardar_gold(self, df_empleados, df_practicantes, carpeta_trabajo):
        """Guarda Gold con versionamiento dual (actual + histórico)"""
        carpeta_gold = carpeta_trabajo / "gold"
        carpeta_gold.mkdir(exist_ok=True)
        
        carpeta_historico = carpeta_gold / "historico"
        carpeta_historico.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime("%d.%m.%Y_%H.%M.%S")
        
        # EMPLEADOS
        if df_empleados.height > 0:
            # Actual
            ruta_emp_actual = carpeta_gold / "bd_empleados_gold.parquet"
            df_empleados.write_parquet(ruta_emp_actual, compression="snappy")
            self.logger.info(f"  ✓ Empleados (actual): {ruta_emp_actual.name}")
            
            ruta_emp_excel = carpeta_gold / "bd_empleados_gold.xlsx"
            df_empleados.write_excel(ruta_emp_excel)
            
            # Histórico
            ruta_emp_hist = carpeta_historico / f"bd_empleados_gold_{timestamp}.parquet"
            df_empleados.write_parquet(ruta_emp_hist, compression="snappy")
            self.logger.info(f"  ✓ Empleados (histórico): {ruta_emp_hist.name}")
        
        # PRACTICANTES
        if df_practicantes.height > 0:
            # Actual
            ruta_prac_actual = carpeta_gold / "bd_practicantes_gold.parquet"
            df_practicantes.write_parquet(ruta_prac_actual, compression="snappy")
            self.logger.info(f"  ✓ Practicantes (actual): {ruta_prac_actual.name}")
            
            ruta_prac_excel = carpeta_gold / "bd_practicantes_gold.xlsx"
            df_practicantes.write_excel(ruta_prac_excel)
            
            # Histórico
            ruta_prac_hist = carpeta_historico / f"bd_practicantes_gold_{timestamp}.parquet"
            df_practicantes.write_parquet(ruta_prac_hist, compression="snappy")
            self.logger.info(f"  ✓ Practicantes (histórico): {ruta_prac_hist.name}")
    
    # ========================================================================
    # STEP 3: APLICACIÓN DE FLAGS
    # ========================================================================
    
    def _step3_aplicar_flags_automatico(self, ruta_cc_actual):
        """
        Aplica flags de negocio a empleados
        Solo necesita bd_empleados_gold.parquet, NO usa CC
        
        CORRECCIÓN: Ahora guarda en /actual + /historico
        """
        tiempo_inicio = time.time()
        
        self.logger.info("")
        self.logger.info("="*70)
        self.logger.info("STEP 3: APLICACIÓN DE FLAGS")
        self.logger.info("="*70)
        
        import polars as pl
        
        # Cargar DuckDB (lazy: solo se importa aquí, no al inicio)
        if not self.duckdb:
            self.logger.info("📦 Cargando DuckDB...")
            try:
                import duckdb
                self.duckdb = duckdb
                self.logger.info("✓ DuckDB cargado")
            except ImportError:
                raise ImportError("DuckDB no está instalado. Instala con: pip install duckdb")
        
        # ✅ CORRECCIÓN: Usar self.output_dir en lugar de calcular desde ruta_cc_actual
        ruta_empleados = self.output_dir / "gold" / "bd_empleados_gold.parquet"
        
        if not ruta_empleados.exists():
            raise FileNotFoundError(f"No se encontró {ruta_empleados}")
        
        # Cargar datos de empleados
        self.logger.info("📊 Cargando datos de empleados...")
        df_empleados = pl.read_parquet(ruta_empleados)
        self.logger.info(f"  ✓ Empleados: {len(df_empleados):,} registros")
        
        # Registrar tabla en DuckDB
        self.logger.info("🦆 Registrando tabla en DuckDB...")
        conn = self.duckdb.connect()
        conn.execute("CREATE OR REPLACE TABLE empleados AS SELECT * FROM df_empleados")
        
        # Cargar query SQL
        self.logger.info("📜 Cargando query SQL...")
        ruta_query_flags = self._buscar_query('queries_flags_gold.sql')
        
        if not ruta_query_flags:
            raise FileNotFoundError("No se encontró queries_flags_gold.sql")
        
        with open(ruta_query_flags, 'r', encoding='utf-8') as f:
            query_flags = f.read()
        
        # Ejecutar query
        self.logger.info("⚙️  Aplicando flags de negocio...")
        result = conn.execute(query_flags).pl()
        
        conn.close()
        
        self.logger.info(f"  ✓ Flags aplicados: {len(result):,} registros")
        
        # ========================================================================
        # ✅ CORRECCIÓN: Implementar patrón dual /actual + /historico
        # ========================================================================
        
        carpeta_gold = self.output_dir / "gold"
        carpeta_historico = carpeta_gold / "historico"
        carpeta_historico.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%d.%m.%Y_%H.%M.%S")
        
        # === ARCHIVOS ACTUALES (sin timestamp) ===
        nombre_actual = "bd_empleados_flags_gold"
        ruta_parquet_actual = carpeta_gold / f"{nombre_actual}.parquet"
        ruta_excel_actual = carpeta_gold / f"{nombre_actual}.xlsx"
        
        self.logger.info("\n💾 Guardando archivos actuales...")
        result.write_parquet(ruta_parquet_actual, compression="snappy")
        self.logger.info(f"  ✓ Parquet (actual): {ruta_parquet_actual.name}")
        
        result.write_excel(ruta_excel_actual)
        self.logger.info(f"  ✓ Excel (actual): {ruta_excel_actual.name}")
        
        # === ARCHIVOS HISTÓRICOS (con timestamp) ===
        nombre_historico = f"bd_empleados_flags_gold_{timestamp}"
        ruta_parquet_historico = carpeta_historico / f"{nombre_historico}.parquet"
        ruta_excel_historico = carpeta_historico / f"{nombre_historico}.xlsx"
        
        self.logger.info("\n📦 Guardando archivos históricos...")
        result.write_parquet(ruta_parquet_historico, compression="snappy")
        self.logger.info(f"  ✓ Parquet (histórico): {ruta_parquet_historico.name}")
        
        result.write_excel(ruta_excel_historico)
        self.logger.info(f"  ✓ Excel (histórico): {ruta_excel_historico.name}")
        
        self.logger.info(f"\n📁 Ubicación:")
        self.logger.info(f"  • Actuales: gold/")
        self.logger.info(f"  • Históricos: gold/historico/")
        
        self.timers['step3'] = time.time() - tiempo_inicio
        
        self.logger.info("-"*70)
        self.logger.info(f"✓ Step 3 completado")
        self.logger.info(f"  • Registros con flags: {len(result):,}")
        self.logger.info(f"  • Archivo actual: {nombre_actual}.parquet")
        self.logger.info(f"  • Archivo histórico: {nombre_historico}.parquet")
        self.logger.info(f"  ⏱️  Duración: {self.logger.format_duration(self.timers['step3'])}")
        self.logger.info("-"*70)
        
        return {
            'registros': len(result),
            'archivo_actual': nombre_actual,
            'archivo_historico': nombre_historico,
            'duracion': self.timers['step3']
        }
    
    # ========================================================================
    # UTILIDADES
    # ========================================================================
    
    def _buscar_esquema(self, nombre_archivo):
        """Busca archivo de esquema usando get_resource_path (compatible con PyInstaller)"""
        from utils.paths import get_resource_path
        
        # Usar get_resource_path que maneja correctamente dev y PyInstaller
        ruta_esquema = get_resource_path(f"esquemas/{nombre_archivo}")
        
        if ruta_esquema.exists():
            return ruta_esquema
        
        self.logger.warning(f"⚠️ Esquema no encontrado: {nombre_archivo}")
        self.logger.warning(f"   Ruta buscada: {ruta_esquema}")
        return None
    
    def _buscar_query(self, nombre_archivo):
        """Busca archivo SQL usando get_resource_path (compatible con PyInstaller)"""
        from utils.paths import get_resource_path
        
        # Usar get_resource_path que maneja correctamente dev y PyInstaller
        ruta_query = get_resource_path(f"queries/{nombre_archivo}")
        
        if ruta_query.exists():
            return ruta_query
        
        self.logger.warning(f"⚠️ Query no encontrado: {nombre_archivo}")
        self.logger.warning(f"   Ruta buscada: {ruta_query}")
        return None