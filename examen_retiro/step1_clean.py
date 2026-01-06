"""
ETL para Programación de Exámenes de Retiro
Procesa archivo Excel con información de empleados cesados y sus exámenes médicos

Arquitectura:
- Bronze: Excel con hoja DATA, headers en fila 3, datos desde fila 4
- Silver: Parquet limpio y estandarizado (SE SOBRESCRIBE EN CADA EJECUCIÓN)
- Output: Excel para visualización de usuario (SE SOBRESCRIBE EN CADA EJECUCIÓN)
"""

import polars as pl
import openpyxl
from pathlib import Path
from datetime import datetime
import time
import sys


def extraer_bronze_examenes_retiro(ruta_excel: str) -> pl.DataFrame:
    """
    Extrae datos de la hoja DATA del archivo Excel
    Headers en fila 3, datos desde fila 4
    """
    print("📂 Extrayendo datos de la capa Bronze...")
    
    try:
        import pandas as pd
        from datetime import datetime, date
        
        # Cargar workbook con openpyxl para extraer headers correctamente
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
        ws = wb['DATA']
        
        # Extraer headers de la fila 3
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=3, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())
            else:
                break  # Detenerse cuando no haya más headers
        
        print(f"   ✓ {len(headers)} columnas identificadas")
        
        # Extraer datos desde fila 4
        data = []
        for row in range(4, ws.max_row + 1):
            first_cell = ws.cell(row=row, column=1).value
            
            # Si la primera celda está vacía, asumir que no hay más datos
            if first_cell is None or str(first_cell).strip() == '':
                break
            
            row_data = []
            for col_idx in range(1, len(headers) + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                
                # Convertir #N/A y errores de Excel a None
                if isinstance(cell_value, str) and cell_value.startswith('#'):
                    cell_value = None
                
                # Mantener datetime/date tal cual, no convertir a string
                row_data.append(cell_value)
            
            data.append(row_data)
        
        wb.close()
        
        print(f"   ✓ {len(data)} filas de datos extraídas")
        
        # Convertir a diccionario de listas
        # Convertir datetime a string ISO para evitar problemas de tipos mixtos
        data_dict = {header: [] for header in headers}
        
        for row_data in data:
            for idx, (header, value) in enumerate(zip(headers, row_data)):
                # Convertir datetime a string ISO para preservar la fecha
                if isinstance(value, (datetime, date)):
                    value = value.strftime('%Y-%m-%d') if isinstance(value, date) else value.strftime('%Y-%m-%d %H:%M:%S')
                # Convertir números a string para evitar problemas con DNI
                elif isinstance(value, (int, float)):
                    value = str(int(value)) if isinstance(value, float) and value == int(value) else str(value)
                
                data_dict[header].append(value)
        
        # Crear DataFrame de Polars desde diccionario
        df = pl.DataFrame(data_dict)
        
        return df
        
    except Exception as e:
        print(f"❌ Error al extraer datos Bronze: {str(e)}")
        raise


def limpiar_silver_examenes_retiro(df_bronze: pl.DataFrame) -> pl.DataFrame:
    """
    Limpia y estandariza los datos para la capa Silver
    """
    print("\n🧹 Procesando capa Silver...")
    
    try:
        df_silver = df_bronze.clone()
        
        # Limpiar espacios en blanco de todas las columnas de texto
        for col in df_silver.columns:
            if df_silver[col].dtype == pl.Utf8:
                df_silver = df_silver.with_columns(
                    pl.col(col).str.strip_chars().alias(col)
                )
        
        # Convertir campos de fecha
        columnas_fecha = ['F. NACIMIENTO', 'FECHA DE CESE', 
                         'UTLIMA FECHA PARA EXAMEN MEDICO', 
                         '2DA CONVOCATORIA EXAMEN MEDICO']
        
        for col in columnas_fecha:
            if col in df_silver.columns:
                # Intentar conversión de fecha
                try:
                    # Primero intentar como datetime
                    df_silver = df_silver.with_columns(
                        pl.col(col).cast(pl.Date, strict=False).alias(col)
                    )
                except:
                    # Si falla, intentar como string y luego parsear
                    df_silver = df_silver.with_columns(
                        pl.col(col).cast(pl.Utf8, strict=False).alias(col)
                    )
        
        # Limpiar columna DNI (convertir a string y quitar decimales)
        if 'DNI' in df_silver.columns:
            df_silver = df_silver.with_columns(
                pl.col('DNI').cast(pl.Utf8, strict=False).str.replace(r'\.0$', '').alias('DNI')
            )
        
        # Reemplazar valores None en columnas de texto con string vacío
        for col in df_silver.columns:
            if df_silver[col].dtype == pl.Utf8:
                df_silver = df_silver.with_columns(
                    pl.col(col).fill_null('').alias(col)
                )
        
        print(f"   ✓ {df_silver.height} registros procesados")
        print(f"   ✓ {df_silver.width} columnas estandarizadas")
        
        return df_silver
        
    except Exception as e:
        print(f"❌ Error en procesamiento Silver: {str(e)}")
        raise


def guardar_silver_parquet(df_silver: pl.DataFrame, ruta_salida: str) -> None:
    """
    Guarda el DataFrame Silver en formato Parquet (SOBRESCRIBE archivo existente)
    """
    print(f"\n💾 Guardando capa Silver (Parquet)...")
    
    try:
        df_silver.write_parquet(ruta_salida)
        
        # Verificar tamaño del archivo
        tamanio_mb = Path(ruta_salida).stat().st_size / (1024 * 1024)
        print(f"   ✓ Archivo guardado: {Path(ruta_salida).name}")
        print(f"   ✓ Tamaño: {tamanio_mb:.2f} MB")
        print(f"   ℹ️  Archivo sobrescrito (sin historial)")
        
    except Exception as e:
        print(f"❌ Error al guardar Parquet: {str(e)}")
        raise


def exportar_excel_usuario(df_silver: pl.DataFrame, ruta_salida: str) -> None:
    """
    Exporta el DataFrame Silver a Excel para visualización de usuario (SOBRESCRIBE archivo existente)
    """
    print(f"\n📊 Exportando Excel para usuario...")
    
    try:
        # Convertir fechas a string para Excel
        df_export = df_silver.clone()
        
        for col in df_export.columns:
            if df_export[col].dtype == pl.Date:
                df_export = df_export.with_columns(
                    pl.col(col).cast(pl.Utf8, strict=False).alias(col)
                )
        
        # Exportar a Excel
        df_export.write_excel(ruta_salida)
        
        print(f"   ✓ Excel guardado: {Path(ruta_salida).name}")
        print(f"   ✓ {df_export.height} filas exportadas")
        print(f"   ℹ️  Archivo sobrescrito (sin historial)")
        
    except Exception as e:
        print(f"❌ Error al exportar Excel: {str(e)}")
        raise


def ejecutar_etl_examenes_retiro(ruta_bronze: str, 
                                   carpeta_silver: str = None) -> None:
    """
    Ejecuta el pipeline ETL completo para exámenes de retiro
    SIN TIMESTAMP - SOBRESCRIBE archivos existentes
    """
    inicio = time.time()
    
    print("=" * 80)
    print("ETL - PROGRAMACIÓN DE EXÁMENES DE RETIRO")
    print("=" * 80)
    print(f"Inicio: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    try:
        # Definir ruta de salida
        if carpeta_silver is None:
            carpeta_silver = Path(ruta_bronze).parent / 'silver'
        
        # Crear carpeta si no existe
        Path(carpeta_silver).mkdir(parents=True, exist_ok=True)
        
        # NOMBRES FIJOS SIN TIMESTAMP (se sobrescriben en cada ejecución)
        ruta_parquet = Path(carpeta_silver) / "examenes_retiro.parquet"
        ruta_excel_output = Path(carpeta_silver) / "examenes_retiro.xlsx"
        
        # 1. Extraer Bronze
        df_bronze = extraer_bronze_examenes_retiro(ruta_bronze)
        
        # 2. Limpiar Silver
        df_silver = limpiar_silver_examenes_retiro(df_bronze)
        
        # 3. Guardar Parquet (sobrescribe)
        guardar_silver_parquet(df_silver, str(ruta_parquet))
        
        # 4. Exportar Excel (sobrescribe)
        exportar_excel_usuario(df_silver, str(ruta_excel_output))
        
        # Resumen final
        duracion = time.time() - inicio
        print("\n" + "=" * 80)
        print("✅ ETL COMPLETADO EXITOSAMENTE")
        print("=" * 80)
        print(f"Registros procesados: {df_silver.height}")
        print(f"Columnas: {df_silver.width}")
        print(f"Tiempo de ejecución: {duracion:.2f} segundos")
        print(f"\nArchivos generados en: {carpeta_silver}")
        print(f"  📦 Parquet: {ruta_parquet.name}")
        print(f"  📊 Excel:   {ruta_excel_output.name}")
        print("\n⚠️  NOTA: Los archivos se sobrescriben en cada ejecución (sin historial)")
        print("=" * 80)
        
    except Exception as e:
        print(f"\n❌ ERROR EN ETL: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    import tkinter as tk
    from tkinter import filedialog
    
    # Ocultar ventana principal de tkinter
    root = tk.Tk()
    root.withdraw()
    
    print("=" * 80)
    print("ETL - PROGRAMACIÓN DE EXÁMENES DE RETIRO")
    print("=" * 80)
    print("\n📂 Seleccione el archivo Excel de exámenes de retiro...\n")
    
    # Abrir explorador de archivos
    ruta_bronze = filedialog.askopenfilename(
        title="Seleccionar archivo Excel - Exámenes de Retiro",
        filetypes=[
            ("Archivos Excel", "*.xlsx *.xlsm *.xls"),
            ("Todos los archivos", "*.*")
        ]
    )
    
    # Verificar si se seleccionó un archivo
    if not ruta_bronze:
        print("❌ No se seleccionó ningún archivo. Proceso cancelado.")
        sys.exit(0)
    
    print(f"✓ Archivo seleccionado: {Path(ruta_bronze).name}\n")
    
    # Definir carpeta silver en la misma ubicación del archivo bronze
    carpeta_bronze = Path(ruta_bronze).parent
    carpeta_silver = carpeta_bronze / 'silver'
    
    print(f"📁 Estructura de salida:")
    print(f"   📂 Archivo Bronze: {carpeta_bronze}")
    print(f"   📦 Carpeta Silver: {carpeta_silver} (parquet + excel)")
    print(f"   ⚠️  Modo: SOBRESCRITURA (sin historial)\n")
    
    # Ejecutar ETL
    ejecutar_etl_examenes_retiro(ruta_bronze, str(carpeta_silver))