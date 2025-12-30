#!/usr/bin/env python3
"""
Script de procesamiento: Excel Bronze → Parquet Silver (Tipo 2)
Características del reporte tipo 2:
- Encabezados en fila 10
- Datos desde fila 11
- Identificación de última fila mediante columna NUMERO DE DOC
"""

import polars as pl
import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import time
import tkinter as tk
from tkinter import filedialog


def select_file(title: str = "Seleccionar archivo Excel") -> Path | None:
    """Abre un explorador de archivos para seleccionar un archivo Excel."""
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=[
            ("Excel files", "*.xlsx *.xlsm *.xls"),
            ("All files", "*.*")
        ]
    )
    
    root.destroy()
    
    return Path(file_path) if file_path else None


def extract_data_from_excel(file_path: Path, header_row: int = 10, data_start_row: int = 11) -> tuple[list[str], list[list]]:
    """
    Extrae encabezados y datos del archivo Excel usando openpyxl.
    Convierte strings en formato DD/MM/YYYY a datetime para estandarización.
    
    Args:
        file_path: Ruta del archivo Excel
        header_row: Número de fila donde están los encabezados (1-indexed)
        data_start_row: Número de fila donde inician los datos (1-indexed)
    
    Returns:
        Tupla con (lista de encabezados, lista de filas de datos)
    """
    from datetime import datetime as dt
    import re
    
    print(f"\n📂 Cargando archivo: {file_path.name}")
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    
    print(f"📄 Hoja activa: {ws.title}")
    
    # Extraer encabezados de la fila especificada
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col_idx).value
        if cell_value:
            headers.append(str(cell_value))
        else:
            # Si encontramos una celda vacía, dejamos de buscar encabezados
            break
    
    print(f"📊 Encabezados detectados: {len(headers)}")
    
    # Buscar la columna NUMERO DE DOC
    numero_doc_col_idx = None
    for idx, header in enumerate(headers, 1):
        if "NUMERO" in header.upper() and "DOC" in header.upper():
            numero_doc_col_idx = idx
            print(f"🔍 Columna 'NUMERO DE DOC' encontrada en posición {idx}: {header}")
            break
    
    if not numero_doc_col_idx:
        raise ValueError("No se encontró la columna 'NUMERO DE DOC' en los encabezados")
    
    # Identificar última fila con datos en NUMERO DE DOC
    last_data_row = data_start_row - 1
    for row_idx in range(data_start_row, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=numero_doc_col_idx).value
        if cell_value is not None and str(cell_value).strip() != "":
            last_data_row = row_idx
    
    total_rows = last_data_row - data_start_row + 1
    print(f"📈 Filas con datos: {total_rows} (desde fila {data_start_row} hasta fila {last_data_row})")
    
    # Patrón para detectar fechas en formato DD/MM/YYYY o D/M/YYYY
    date_pattern = re.compile(r'^\d{1,2}/\d{1,2}/\d{4}')
    
    # Contador de conversiones
    conversiones_fecha = 0
    
    # Extraer datos
    data_rows = []
    for row_idx in range(data_start_row, last_data_row + 1):
        row_data = []
        for col_idx in range(1, len(headers) + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            
            # Si el valor es un string que parece una fecha en formato DD/MM/YYYY
            if isinstance(cell_value, str):
                # Limpiar tabs y espacios
                cell_value = cell_value.strip()
                
                if date_pattern.match(cell_value):
                    try:
                        # Intentar convertir DD/MM/YYYY a datetime
                        cell_value = dt.strptime(cell_value, "%d/%m/%Y")
                        conversiones_fecha += 1
                    except ValueError:
                        # Si falla, dejar el valor original
                        pass
            
            row_data.append(cell_value)
        data_rows.append(row_data)
    
    if conversiones_fecha > 0:
        print(f"🔧 Convertidas {conversiones_fecha} fechas de formato DD/MM/YYYY a datetime")
    
    wb.close()
    
    return headers, data_rows


def create_polars_dataframe(headers: list[str], data_rows: list[list]) -> pl.DataFrame:
    """
    Crea un DataFrame de Polars a partir de encabezados y datos.
    Usa pandas como intermediario para manejar mejor los tipos mixtos.
    Convierte datetime a string con formato consistente.
    
    Args:
        headers: Lista de nombres de columnas
        data_rows: Lista de filas de datos
    
    Returns:
        DataFrame de Polars
    """
    import pandas as pd
    from datetime import datetime as dt
    
    # Procesar data_rows para convertir datetime a string con formato consistente
    processed_rows = []
    for row in data_rows:
        processed_row = []
        for value in row:
            # Convertir datetime a string con formato consistente YYYY-MM-DD HH:MM:SS
            if isinstance(value, dt):
                processed_row.append(value.strftime("%Y-%m-%d %H:%M:%S"))
            else:
                processed_row.append(value)
        processed_rows.append(processed_row)
    
    # Crear DataFrame de pandas con los datos procesados
    df_pandas = pd.DataFrame(processed_rows, columns=headers)
    
    # Convertir todas las columnas a string para evitar conflictos de tipo
    df_pandas = df_pandas.astype(str)
    
    # Reemplazar 'None' y 'nan' strings por valores nulos reales
    df_pandas = df_pandas.replace(['None', 'nan', 'NaT'], None)
    
    # Convertir a Polars
    df = pl.from_pandas(df_pandas)
    
    return df


def save_to_parquet(df: pl.DataFrame, output_path: Path) -> None:
    """Guarda el DataFrame en formato Parquet."""
    print(f"\n💾 Guardando archivo Parquet: {output_path.name}")
    df.write_parquet(output_path)
    print(f"✓ Archivo Parquet guardado: {output_path}")


def save_to_excel_visualization(df: pl.DataFrame, output_path: Path) -> None:
    """Guarda el DataFrame en formato Excel para visualización."""
    print(f"\n📊 Generando archivo Excel de visualización: {output_path.name}")
    
    # Convertir a pandas para usar to_excel con formato
    df_pandas = df.to_pandas()
    
    # Crear archivo Excel con formato
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_pandas.to_excel(writer, sheet_name='Datos', index=False)
        
        # Obtener la hoja y aplicar formato
        worksheet = writer.sheets['Datos']
        
        # Formato de encabezado
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, col_name in enumerate(df_pandas.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✓ Archivo Excel de visualización guardado: {output_path}")


def main():
    """Función principal del script."""
    print("=" * 70)
    print("  PROCESAMIENTO: EXCEL BRONZE → PARQUET SILVER (TIPO 2)")
    print("  Encabezados en fila 10 | Datos desde fila 11")
    print("=" * 70)
    
    # Iniciar temporizador
    start_time = time.time()
    
    # Paso 1: Seleccionar archivo Excel de entrada
    print("\n📁 Selecciona el archivo Excel Bronze...")
    input_file = select_file("Seleccionar archivo Excel Bronze (Tipo 2)")
    
    if not input_file:
        print("\n❌ No se seleccionó ningún archivo. Proceso cancelado.")
        return
    
    if not input_file.exists():
        print(f"\n❌ El archivo no existe: {input_file}")
        return
    
    # Paso 2: Extraer datos del Excel
    try:
        headers, data_rows = extract_data_from_excel(
            input_file,
            header_row=10,
            data_start_row=11
        )
    except Exception as e:
        print(f"\n❌ Error al extraer datos del Excel: {e}")
        return
    
    # Paso 3: Crear DataFrame de Polars
    print(f"\n🔄 Creando DataFrame de Polars...")
    df = create_polars_dataframe(headers, data_rows)
    print(f"✓ DataFrame creado: {df.shape[0]} filas × {df.shape[1]} columnas")
    
    # Paso 4: Generar nombres de archivo de salida
    # Los archivos silver se guardan en la misma ubicación del Excel de entrada
    timestamp = datetime.now().strftime("%d.%m.%Y_%H.%M.%S")
    output_dir = input_file.parent  # Mismo directorio del archivo de entrada
    
    parquet_filename = f"BD_{timestamp}_silver.parquet"
    excel_filename = f"BD_{timestamp}_silver.xlsx"
    
    parquet_path = output_dir / parquet_filename
    excel_path = output_dir / excel_filename
    
    # Paso 5: Guardar Parquet
    try:
        save_to_parquet(df, parquet_path)
    except Exception as e:
        print(f"\n❌ Error al guardar archivo Parquet: {e}")
        return
    
    # Paso 6: Guardar Excel de visualización
    try:
        import pandas as pd
        save_to_excel_visualization(df, excel_path)
    except Exception as e:
        print(f"\n❌ Error al guardar archivo Excel: {e}")
        return
    
    # Calcular tiempo de ejecución
    elapsed_time = time.time() - start_time
    
    # Resumen final
    print("\n" + "=" * 70)
    print("✅ PROCESO COMPLETADO EXITOSAMENTE")
    print("=" * 70)
    print(f"\n📊 Estadísticas:")
    print(f"   • Filas procesadas: {df.shape[0]:,}")
    print(f"   • Columnas: {df.shape[1]}")
    print(f"   • Tiempo de ejecución: {elapsed_time:.2f} segundos")
    
    print(f"\n📁 Archivos generados:")
    print(f"   • Parquet: {parquet_path.name}")
    print(f"   • Excel:   {excel_path.name}")
    print(f"\n📂 Ubicación: {output_dir}")
    print("=" * 70)


if __name__ == "__main__":
    main()