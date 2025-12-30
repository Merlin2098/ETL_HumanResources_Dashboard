#!/usr/bin/env python3
"""
Script de procesamiento: Parquet Silver → Parquet + Excel Gold (Tipo 2)
Valida columnas según esquema JSON y genera capa Gold filtrada
"""

import polars as pl
import pandas as pd
from pathlib import Path
from datetime import datetime
import json
import time
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import Font, PatternFill, Alignment


def select_file(title: str = "Seleccionar archivo Parquet Silver") -> Path | None:
    """Abre un explorador de archivos para seleccionar un archivo Parquet."""
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=[
            ("Parquet files", "*.parquet"),
            ("All files", "*.*")
        ]
    )
    
    root.destroy()
    
    return Path(file_path) if file_path else None


def find_esquemas_folder() -> Path | None:
    """
    Busca la carpeta 'esquemas' en el directorio actual y hasta 3 niveles arriba.
    
    Returns:
        Path de la carpeta esquemas o None si no se encuentra
    """
    carpeta_actual = Path.cwd()
    
    # Buscar en el directorio actual y hasta 3 niveles arriba
    for _ in range(4):
        posible_esquemas = carpeta_actual / "esquemas"
        if posible_esquemas.exists() and posible_esquemas.is_dir():
            return posible_esquemas
        carpeta_actual = carpeta_actual.parent
    
    return None


def load_schema() -> dict:
    """
    Busca la carpeta 'esquemas' y abre un explorador de archivos para que el usuario
    seleccione el esquema JSON.
    
    Returns:
        Diccionario con el esquema seleccionado
    """
    # Buscar carpeta de esquemas
    carpeta_esquemas = find_esquemas_folder()
    
    if carpeta_esquemas is None:
        raise FileNotFoundError(
            f"No se encontró la carpeta 'esquemas' en el proyecto.\n"
            f"Asegúrate de que exista la carpeta 'esquemas' en la raíz del proyecto."
        )
    
    # Listar archivos JSON disponibles
    esquemas_disponibles = list(carpeta_esquemas.glob("*.json"))
    
    if not esquemas_disponibles:
        raise FileNotFoundError(
            f"No se encontraron archivos JSON en: {carpeta_esquemas}\n"
            f"Por favor, coloca los archivos de esquemas (.json) en esta carpeta."
        )
    
    print(f"📁 Carpeta de esquemas: {carpeta_esquemas}")
    print(f"✓ Esquemas disponibles:")
    for i, esquema in enumerate(esquemas_disponibles, 1):
        print(f"   {i}. {esquema.name}")
    print()
    
    # Abrir explorador para seleccionar el esquema JSON
    print("📁 Seleccione el archivo JSON del esquema Gold...")
    
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    ruta_schema = filedialog.askopenfilename(
        title="Seleccione el esquema JSON Gold",
        initialdir=str(carpeta_esquemas),
        filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
    )
    root.destroy()
    
    if not ruta_schema:
        raise ValueError("No se seleccionó archivo de esquema. Operación cancelada.")
    
    schema_path = Path(ruta_schema)
    print(f"✓ Esquema seleccionado: {schema_path.name}")
    print(f"📋 Cargando esquema desde: {schema_path}")
    
    with open(schema_path, 'r', encoding='utf-8') as f:
        schema = json.load(f)
    
    return schema


def validate_and_filter_columns(df: pl.DataFrame, schema: dict) -> tuple[pl.DataFrame, dict]:
    """
    Valida que las columnas del esquema existan en el DataFrame y filtra solo esas columnas.
    PRESERVA EL ORDEN ORIGINAL del DataFrame Silver.
    
    Args:
        df: DataFrame de Polars
        schema: Diccionario con el esquema
    
    Returns:
        Tupla con (DataFrame filtrado, estadísticas de validación)
    """
    required_columns = [col["name"] for col in schema["columns"]]
    present_columns = df.columns
    
    # Verificar columnas presentes y faltantes
    present = [col for col in required_columns if col in present_columns]
    missing = [col for col in required_columns if col not in present_columns]
    
    # IMPORTANTE: Filtrar columnas en el ORDEN ORIGINAL del DataFrame Silver
    # No usar el orden del schema
    columns_to_select = [col for col in df.columns if col in required_columns]
    
    stats = {
        "total_required": len(required_columns),
        "present": len(present),
        "missing": len(missing),
        "missing_list": missing,
        "present_list": present
    }
    
    if missing:
        print(f"\n⚠️  ADVERTENCIA: {len(missing)} columna(s) faltante(s):")
        for col in missing:
            print(f"   • {col}")
    
    # Filtrar solo las columnas presentes del esquema, pero en el orden original
    df_filtered = df.select(columns_to_select)
    
    return df_filtered, stats


def convert_date_columns(df: pl.DataFrame, schema: dict) -> pl.DataFrame:
    """
    Convierte columnas de tipo date según el esquema.
    Asume que Silver ya tiene fechas estandarizadas en formato YYYY-MM-DD HH:MM:SS
    
    Args:
        df: DataFrame de Polars
        schema: Diccionario con el esquema
    
    Returns:
        DataFrame con columnas de fecha convertidas
    """
    date_columns = [col["name"] for col in schema["columns"] if col["type"] == "date"]
    
    for col_name in date_columns:
        if col_name in df.columns:
            # Intentar convertir con formato completo YYYY-MM-DD HH:MM:SS
            try:
                df = df.with_columns(
                    pl.col(col_name).str.to_date(format="%Y-%m-%d %H:%M:%S", strict=False)
                    .alias(col_name)
                )
                print(f"✓ Columna '{col_name}' convertida a fecha")
            except Exception as e:
                print(f"⚠️  No se pudo convertir '{col_name}' a fecha: {e}")
                # Intentar formato alternativo sin hora
                try:
                    df = df.with_columns(
                        pl.col(col_name).str.to_date(format="%Y-%m-%d", strict=False)
                        .alias(col_name)
                    )
                    print(f"✓ Columna '{col_name}' convertida a fecha (formato alternativo)")
                except:
                    print(f"⚠️  '{col_name}' permanece como string")
    
    return df


def save_to_parquet(df: pl.DataFrame, output_path: Path) -> None:
    """Guarda el DataFrame en formato Parquet."""
    print(f"\n💾 Guardando archivo Parquet Gold: {output_path.name}")
    df.write_parquet(output_path)
    print(f"✓ Archivo Parquet guardado: {output_path}")


def save_to_excel(df: pl.DataFrame, output_path: Path) -> None:
    """Guarda el DataFrame en formato Excel con formato profesional."""
    print(f"\n📊 Generando archivo Excel Gold: {output_path.name}")
    
    # Convertir a pandas
    df_pandas = df.to_pandas()
    
    # Crear archivo Excel con formato
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_pandas.to_excel(writer, sheet_name='Datos', index=False)
        
        # Obtener la hoja y aplicar formato
        worksheet = writer.sheets['Datos']
        
        # Formato de encabezado
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, col_name in enumerate(df_pandas.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    print(f"✓ Archivo Excel guardado: {output_path}")


def main():
    """Función principal del script."""
    print("=" * 70)
    print("  PROCESAMIENTO: PARQUET SILVER → PARQUET + EXCEL GOLD (TIPO 2)")
    print("=" * 70)
    
    # Iniciar temporizador
    start_time = time.time()
    
    # Paso 1: Seleccionar archivo Parquet Silver
    print("\n📁 Selecciona el archivo Parquet Silver...")
    input_file = select_file("Seleccionar archivo Parquet Silver (Tipo 2)")
    
    if not input_file:
        print("\n❌ No se seleccionó ningún archivo. Proceso cancelado.")
        return
    
    if not input_file.exists():
        print(f"\n❌ El archivo no existe: {input_file}")
        return
    
    # Paso 2: Cargar esquema JSON
    try:
        schema = load_schema()
        print(f"✓ Esquema cargado: {schema['schema_name']} v{schema['version']}")
        print(f"   Columnas esperadas: {len(schema['columns'])}")
    except Exception as e:
        print(f"\n❌ Error al cargar esquema: {e}")
        return
    
    # Paso 3: Cargar Parquet Silver
    print(f"\n📂 Cargando Parquet Silver: {input_file.name}")
    try:
        df = pl.read_parquet(input_file)
        print(f"✓ Datos cargados: {df.shape[0]} filas × {df.shape[1]} columnas")
    except Exception as e:
        print(f"\n❌ Error al cargar Parquet: {e}")
        return
    
    # Paso 4: Validar y filtrar columnas
    print(f"\n🔍 Validando columnas según esquema...")
    df_gold, stats = validate_and_filter_columns(df, schema)
    
    print(f"\n📊 Resultado de validación:")
    print(f"   • Columnas requeridas: {stats['total_required']}")
    print(f"   • Columnas presentes: {stats['present']}")
    print(f"   • Columnas faltantes: {stats['missing']}")
    
    if stats['missing']:
        print(f"\n⚠️  El proceso continuará sin las columnas faltantes")
    
    print(f"\n✓ DataFrame Gold: {df_gold.shape[0]} filas × {df_gold.shape[1]} columnas")
    
    # Paso 5: Convertir columnas de fecha
    print(f"\n🔄 Convirtiendo columnas de fecha...")
    df_gold = convert_date_columns(df_gold, schema)
    
    # Paso 6: Mostrar tipos de datos finales
    print(f"\n📋 Tipos de datos finales:")
    for col in df_gold.columns:
        print(f"   • {col}: {df_gold[col].dtype}")
    
    # Paso 7: Dividir DataFrame según Modalidad de Contrato
    print(f"\n🔀 Dividiendo datos según Modalidad de Contrato...")
    print("-" * 70)
    
    # Verificar que la columna Modalidad de Contrato existe
    if "Modalidad de Contrato" not in df_gold.columns:
        print(f"❌ La columna 'Modalidad de Contrato' no existe en el DataFrame")
        return
    
    # Dividir en practicantes y empleados
    df_practicantes = df_gold.filter(
        pl.col("Modalidad de Contrato").str.contains("TERMINO DE CONVENIO")
    )
    
    df_empleados = df_gold.filter(
        ~pl.col("Modalidad de Contrato").str.contains("TERMINO DE CONVENIO")
    )
    
    print(f"✓ Practicantes (TERMINO DE CONVENIO): {df_practicantes.shape[0]:,} registros")
    print(f"✓ Empleados (otros): {df_empleados.shape[0]:,} registros")
    print(f"✓ Total: {df_gold.shape[0]:,} registros")
    print("-" * 70)
    
    # Paso 8: Generar nombres de archivo de salida
    # Los archivos gold se guardan en la misma ubicación del parquet silver
    timestamp = datetime.now().strftime("%d.%m.%Y_%H.%M.%S")
    output_dir = input_file.parent
    
    # Archivos para practicantes
    parquet_practicantes = output_dir / f"BD_Practicantes_{timestamp}_gold.parquet"
    excel_practicantes = output_dir / f"BD_Practicantes_{timestamp}_gold.xlsx"
    
    # Archivos para empleados
    parquet_empleados = output_dir / f"BD_Empleados_{timestamp}_gold.parquet"
    excel_empleados = output_dir / f"BD_Empleados_{timestamp}_gold.xlsx"
    
    # Paso 9: Guardar archivos de Practicantes
    if df_practicantes.shape[0] > 0:
        print(f"\n💾 Guardando archivos de PRACTICANTES...")
        try:
            save_to_parquet(df_practicantes, parquet_practicantes)
            save_to_excel(df_practicantes, excel_practicantes)
        except Exception as e:
            print(f"\n❌ Error al guardar archivos de practicantes: {e}")
    else:
        print(f"\n⚠️  No hay registros de practicantes para guardar")
    
    # Paso 10: Guardar archivos de Empleados
    if df_empleados.shape[0] > 0:
        print(f"\n💾 Guardando archivos de EMPLEADOS...")
        try:
            save_to_parquet(df_empleados, parquet_empleados)
            save_to_excel(df_empleados, excel_empleados)
        except Exception as e:
            print(f"\n❌ Error al guardar archivos de empleados: {e}")
    else:
        print(f"\n⚠️  No hay registros de empleados para guardar")
    
    # Calcular tiempo de ejecución
    elapsed_time = time.time() - start_time
    
    # Resumen final
    print("\n" + "=" * 70)
    print("✅ PROCESO COMPLETADO EXITOSAMENTE")
    print("=" * 70)
    print(f"\n📊 Estadísticas:")
    print(f"   • Total de registros procesados: {df_gold.shape[0]:,}")
    print(f"   • Columnas: {df_gold.shape[1]}")
    print(f"   • Practicantes: {df_practicantes.shape[0]:,}")
    print(f"   • Empleados: {df_empleados.shape[0]:,}")
    print(f"   • Tiempo de ejecución: {elapsed_time:.2f} segundos")
    
    print(f"\n📁 Archivos generados:")
    
    if df_practicantes.shape[0] > 0:
        parquet_size_mb = parquet_practicantes.stat().st_size / (1024 * 1024)
        excel_size_mb = excel_practicantes.stat().st_size / (1024 * 1024)
        print(f"\n   PRACTICANTES:")
        print(f"   • Parquet: {parquet_practicantes.name} ({parquet_size_mb:.2f} MB)")
        print(f"   • Excel:   {excel_practicantes.name} ({excel_size_mb:.2f} MB)")
    
    if df_empleados.shape[0] > 0:
        parquet_size_mb = parquet_empleados.stat().st_size / (1024 * 1024)
        excel_size_mb = excel_empleados.stat().st_size / (1024 * 1024)
        print(f"\n   EMPLEADOS:")
        print(f"   • Parquet: {parquet_empleados.name} ({parquet_size_mb:.2f} MB)")
        print(f"   • Excel:   {excel_empleados.name} ({excel_size_mb:.2f} MB)")
    
    print(f"\n📂 Ubicación: {output_dir}")
    print("=" * 70)


if __name__ == "__main__":
    main()