"""
Script: step1_clean.py
Descripci√≥n: Procesa archivo Excel de Programaci√≥n de Ex√°menes de Retiro
             - Hoja DATA: headers en fila 3, datos desde fila 4
             
Arquitectura:
- Bronze: Excel con informaci√≥n de empleados cesados y ex√°menes m√©dicos
- Silver: Parquet limpio y estandarizado (SE SOBRESCRIBE EN CADA EJECUCI√ìN)

Salida: Archivos sin timestamp en carpeta silver/
    - examenes_retiro.parquet
    - examenes_retiro.xlsx

Autor: Richi
Fecha: 06.01.2025
"""

import polars as pl
import openpyxl
from pathlib import Path
from datetime import datetime, date
import time
import sys
from tkinter import Tk, filedialog


def seleccionar_archivo_excel() -> Path | None:
    """Abre di√°logo para seleccionar archivo Excel"""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    archivo = filedialog.askopenfilename(
        title="Seleccionar archivo Excel - Ex√°menes de Retiro (Bronze)",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")]
    )
    
    root.destroy()
    
    return Path(archivo) if archivo else None


def extraer_bronze_examenes_retiro(ruta_excel: Path) -> pl.DataFrame:
    """
    Extrae datos de la hoja DATA del archivo Excel
    Headers en fila 3, datos desde fila 4
    
    OPTIMIZACI√ìN: Usa iter_rows con values_only para mejor rendimiento
    
    Args:
        ruta_excel: Path al archivo Excel Bronze
        
    Returns:
        DataFrame de Polars con los datos extra√≠dos
    """
    print("\nüìÇ Extrayendo datos de la capa Bronze...")
    print(f"   Archivo: {ruta_excel.name}")
    
    try:
        # Cargar workbook con openpyxl
        wb = openpyxl.load_workbook(ruta_excel, data_only=True, read_only=True)
        
        if 'DATA' not in wb.sheetnames:
            raise ValueError(f"La hoja 'DATA' no existe. Hojas disponibles: {wb.sheetnames}")
        
        ws = wb['DATA']
        
        # Extraer headers de la fila 3 (m√°s eficiente con iter_rows)
        headers = []
        for cell in ws[3]:  # Fila 3
            if cell.value is not None:
                headers.append(str(cell.value).strip())
            else:
                break
        
        print(f"   ‚úì {len(headers)} columnas identificadas")
        
        # OPTIMIZACI√ìN: Usar iter_rows con values_only=True para mejor rendimiento
        # Solo iteramos hasta la columna que tiene headers
        data = []
        
        for row in ws.iter_rows(min_row=4, max_col=len(headers), values_only=True):
            # Detener si la primera celda est√° vac√≠a (fin de datos)
            if row[0] is None or (isinstance(row[0], str) and row[0].strip() == ''):
                break
            
            # Convertir fila a lista y procesar valores
            row_data = []
            for cell_value in row:
                # Convertir #N/A y errores de Excel a None
                if isinstance(cell_value, str) and cell_value.startswith('#'):
                    cell_value = None
                # Convertir datetime/date a string ISO
                elif isinstance(cell_value, datetime):
                    cell_value = cell_value.strftime('%Y-%m-%d %H:%M:%S')
                elif isinstance(cell_value, date):
                    cell_value = cell_value.strftime('%Y-%m-%d')
                # Convertir n√∫meros a string para evitar problemas con DNI
                elif isinstance(cell_value, (int, float)):
                    if isinstance(cell_value, float) and cell_value == int(cell_value):
                        cell_value = str(int(cell_value))
                    else:
                        cell_value = str(cell_value)
                
                row_data.append(cell_value)
            
            data.append(row_data)
        
        wb.close()
        
        print(f"   ‚úì {len(data)} filas de datos extra√≠das")
        
        # Crear DataFrame de Polars
        if not data:
            print("   ‚ö†Ô∏è  No se encontraron datos")
            return pl.DataFrame()
        
        # Crear diccionario para DataFrame
        data_dict = {header: [row[i] if i < len(row) else None for row in data] 
                     for i, header in enumerate(headers)}
        
        # Crear DataFrame con strict=False para manejar tipos mixtos
        df = pl.DataFrame(data_dict, strict=False)
        
        return df
        
    except Exception as e:
        print(f"   ‚ùå Error al extraer datos Bronze: {str(e)}")
        raise


def limpiar_silver_examenes_retiro(df_bronze: pl.DataFrame) -> pl.DataFrame:
    """
    Limpia y estandariza los datos para la capa Silver
    
    Args:
        df_bronze: DataFrame extra√≠do de Bronze
        
    Returns:
        DataFrame limpio para Silver
    """
    print("\nüßπ Procesando capa Silver...")
    
    try:
        df_silver = df_bronze.clone()
        
        # Limpiar espacios en blanco de todas las columnas de texto
        for col in df_silver.columns:
            if df_silver[col].dtype == pl.Utf8:
                df_silver = df_silver.with_columns(
                    pl.col(col).str.strip_chars().alias(col)
                )
        
        # Convertir campos de fecha desde string a Date
        columnas_fecha = ['F. NACIMIENTO', 'FECHA DE CESE', 
                         'UTLIMA FECHA PARA EXAMEN MEDICO', 
                         '2DA CONVOCATORIA EXAMEN MEDICO']
        
        for col in columnas_fecha:
            if col in df_silver.columns:
                # Intentar conversi√≥n desde string con timestamp
                try:
                    df_silver = df_silver.with_columns(
                        pl.col(col)
                        .str.to_datetime(format="%Y-%m-%d %H:%M:%S", strict=False)
                        .cast(pl.Date, strict=False)
                        .alias(col)
                    )
                except:
                    # Si falla con timestamp, intentar solo fecha
                    try:
                        df_silver = df_silver.with_columns(
                            pl.col(col)
                            .str.to_datetime(format="%Y-%m-%d", strict=False)
                            .cast(pl.Date, strict=False)
                            .alias(col)
                        )
                    except:
                        # Si todo falla, dejar como string
                        pass
        
        # Limpiar columna DNI (convertir a string y quitar decimales)
        if 'DNI' in df_silver.columns:
            df_silver = df_silver.with_columns(
                pl.col('DNI').cast(pl.Utf8, strict=False).str.replace(r'\.0$', '').alias('DNI')
            )
        
        # Reemplazar valores None en columnas de texto con string vac√≠o
        for col in df_silver.columns:
            if df_silver[col].dtype == pl.Utf8:
                df_silver = df_silver.with_columns(
                    pl.col(col).fill_null('').alias(col)
                )
        
        print(f"   ‚úì {df_silver.height} registros procesados")
        print(f"   ‚úì {df_silver.width} columnas estandarizadas")
        
        return df_silver
        
    except Exception as e:
        print(f"   ‚ùå Error en procesamiento Silver: {str(e)}")
        raise


def guardar_resultados(df_silver: pl.DataFrame, carpeta_trabajo: Path):
    """
    Guarda el DataFrame Silver como parquet y Excel en carpeta silver/
    Sin timestamp - se sobreescribe en cada ejecuci√≥n
    
    Args:
        df_silver: DataFrame a guardar
        carpeta_trabajo: Path de la carpeta de trabajo
        
    Returns:
        tuple: (ruta_parquet, ruta_excel)
    """
    # Crear carpeta silver si no existe
    carpeta_silver = carpeta_trabajo / "silver"
    carpeta_silver.mkdir(exist_ok=True)
    
    # Nombres fijos sin timestamp
    nombre_base = "examenes_retiro"
    ruta_parquet = carpeta_silver / f"{nombre_base}.parquet"
    ruta_excel = carpeta_silver / f"{nombre_base}.xlsx"
    
    print(f"\n[2/2] Guardando resultados en capa Silver...")
    print(f"  üìÅ Carpeta: {carpeta_silver}")
    
    # Guardar Parquet
    print(f"  - Guardando parquet...", end='', flush=True)
    df_silver.write_parquet(ruta_parquet, compression="snappy")
    print(f" ‚úì")
    print(f"    Ubicaci√≥n: {ruta_parquet.name}")
    
    # Guardar Excel
    print(f"  - Guardando Excel...", end='', flush=True)
    # Convertir fechas a string para Excel
    df_export = df_silver.clone()
    for col in df_export.columns:
        if df_export[col].dtype == pl.Date:
            df_export = df_export.with_columns(
                pl.col(col).cast(pl.Utf8, strict=False).alias(col)
            )
    df_export.write_excel(ruta_excel)
    print(f" ‚úì")
    print(f"    Ubicaci√≥n: {ruta_excel.name}")
    
    return ruta_parquet, ruta_excel


def main():
    """Funci√≥n principal de procesamiento"""
    print("=" * 80)
    print(" PROCESADOR DE EX√ÅMENES DE RETIRO - CAPA SILVER ".center(80, "="))
    print("=" * 80)
    
    # 1. Seleccionar archivo
    print("\n[PASO 1] Selecciona el archivo Excel de Ex√°menes de Retiro (Bronze)...")
    archivo_bronze = seleccionar_archivo_excel()
    
    if not archivo_bronze:
        print("‚ùå No se seleccion√≥ ning√∫n archivo. Proceso cancelado.")
        return
    
    # Iniciar cron√≥metro despu√©s de la selecci√≥n
    tiempo_inicio = time.time()
    
    print(f"‚úì Archivo seleccionado: {archivo_bronze.name}")
    carpeta_trabajo = archivo_bronze.parent
    
    # 2. Procesar datos
    print("\n" + "=" * 80)
    print(" PROCESAMIENTO ".center(80, "="))
    print("=" * 80)
    print(f"\n[1/2] Extrayendo y limpiando datos...")
    
    try:
        # 2.1 Extraer Bronze
        df_bronze = extraer_bronze_examenes_retiro(archivo_bronze)
        
        if df_bronze.is_empty():
            print("‚ùå No se encontraron datos en el archivo.")
            return
        
        # 2.2 Limpiar Silver
        df_silver = limpiar_silver_examenes_retiro(df_bronze)
        
        # 3. Guardar resultados
        ruta_parquet, ruta_excel = guardar_resultados(df_silver, carpeta_trabajo)
        
        # Calcular tiempo total
        tiempo_total = time.time() - tiempo_inicio
        
        # 4. Resumen final
        print("\n" + "=" * 80)
        print(" RESUMEN ".center(80, "="))
        print("=" * 80)
        
        print(f"\n‚úì Procesamiento completado exitosamente")
        print(f"\nüìä Estad√≠sticas:")
        print(f"  - Total de registros: {len(df_silver):,}")
        print(f"  - Total de columnas: {len(df_silver.columns)}")
        
        print(f"\nüìÅ Archivos generados en carpeta silver/:")
        print(f"  - Parquet: {ruta_parquet.name}")
        print(f"  - Excel:   {ruta_excel.name}")
        
        print(f"\n‚è±Ô∏è  Tiempo de ejecuci√≥n: {tiempo_total:.2f}s")
        
        print("\nüí° Los archivos se sobreescriben en cada ejecuci√≥n (sin historial)")
        
        print("\n" + "=" * 80)
        
    except Exception as e:
        print(f"\n‚ùå Error durante el procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n‚ùå Error fatal: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)