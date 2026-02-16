"""
Script de consolidación de reportes de planilla
Consolida múltiples archivos Excel en un solo parquet/Excel en capa Silver

REFACTORIZADO para compatibilidad con worker UI:
- consolidar_archivos(): Acepta lista de archivos directamente y GUARDA resultados
- guardar_resultados(): Guarda en carpeta silver
- main(): Mantiene funcionalidad standalone con selección de carpeta
"""

import polars as pl
import re
from pathlib import Path
from datetime import datetime
import time
import traceback


def extraer_periodo(nombre_archivo):
    """
    Extrae el periodo del nombre del archivo
    
    Patrón esperado: METSO_Planilla YYYY-MM Empleados.xlsx
    
    Args:
        nombre_archivo: Nombre del archivo (str)
        
    Returns:
        str: Periodo en formato YYYY-MM o None si no se encuentra
    """
    # Patrón para capturar YYYY-MM
    patron = r'(\d{4}-\d{2})'
    match = re.search(patron, nombre_archivo)
    
    if match:
        return match.group(1)
    else:
        return None


def leer_archivo_planilla(archivo_path, periodo):
    """
    Lee un archivo de planilla y agrega la columna PERIODO
    
    Args:
        archivo_path: Path del archivo Excel
        periodo: Periodo a agregar (str)
        
    Returns:
        pl.DataFrame: DataFrame con los datos del archivo
    """
    try:
        from openpyxl import load_workbook
        
        # Leer con openpyxl para obtener encabezados correctos
        wb = load_workbook(archivo_path, data_only=True, read_only=True)
        
        # Leer específicamente la hoja "Planilla"
        if "Planilla" not in wb.sheetnames:
            raise Exception(f"La hoja 'Planilla' no existe en el archivo. Hojas disponibles: {', '.join(wb.sheetnames)}")
        
        ws = wb["Planilla"]
        
        # Leer encabezados desde fila 6
        encabezados = []
        for col in range(1, 200):  # Leer hasta columna 200
            cell = ws.cell(row=6, column=col)
            valor = cell.value
            
            if valor is not None:
                encabezado = str(valor).strip()
                encabezados.append(encabezado)
            else:
                # Si encontramos 3 columnas vacías consecutivas, terminamos
                if col > len(encabezados) + 3:
                    break
        
        # Leer datos desde fila 7 en adelante
        datos = []
        for row in ws.iter_rows(min_row=7, values_only=True):
            if any(cell is not None for cell in row):
                # Filtrar filas que parecen ser subtítulos o totales
                primera_celda = row[0] if len(row) > 0 else None
                
                # Si la primera celda es texto largo o None, probablemente no es una fila de datos
                if primera_celda is None:
                    continue
                if isinstance(primera_celda, str) and len(primera_celda) > 20:
                    continue
                    
                datos.append(list(row[:len(encabezados)]))
        
        wb.close()
        
        # Crear diccionario para DataFrame
        data_dict = {encabezados[i]: [fila[i] if i < len(fila) else None for fila in datos] 
                     for i in range(len(encabezados))}
        
        # Crear DataFrame con polars usando strict=False para manejar tipos mixtos
        # Intentamos crear el DataFrame de forma más robusta
        try:
            df = pl.DataFrame(data_dict, strict=False)
        except Exception as e:
            # Si falla, convertir todo a string primero
            data_dict_str = {k: [str(v) if v is not None else None for v in valores] 
                           for k, valores in data_dict.items()}
            df = pl.DataFrame(data_dict_str, strict=False)
        
        # Agregar columna PERIODO al inicio
        df = df.with_columns(
            pl.lit(periodo).alias("PERIODO")
        )
        
        # Reorganizar para que PERIODO sea la primera columna
        columnas = ["PERIODO"] + [col for col in df.columns if col != "PERIODO"]
        df = df.select(columnas)
        
        return df
        
    except Exception as e:
        raise Exception(f"Error al leer {archivo_path.name}: {str(e)}")


def consolidar_archivos(archivos, carpeta_trabajo):
    """
    Consolida múltiples archivos de planilla en un solo DataFrame
    y GUARDA los resultados en capa Silver
    
    Args:
        archivos: Lista de Path de archivos Excel O Path de carpeta (str/Path)
        carpeta_trabajo: Path de la carpeta de trabajo para guardar resultados
        
    Returns:
        pl.DataFrame: DataFrame consolidado
    """
    # ====================================================================
    # DETECTAR MODO: Lista de archivos (worker) vs carpeta (standalone)
    # ====================================================================
    
    # Si recibe una lista de archivos (modo worker)
    if isinstance(archivos, list) and len(archivos) > 0 and isinstance(archivos[0], (str, Path)):
        archivos_excel = [Path(f) for f in archivos]
        print(f"\n[MODO WORKER] Procesando {len(archivos_excel)} archivo(s) seleccionado(s)...")
    
    # Si recibe una carpeta (modo standalone o legado)
    elif isinstance(archivos, (str, Path)):
        carpeta = Path(archivos)
        print(f"\n[MODO STANDALONE] Buscando archivos Excel en: {carpeta}")
        
        # Buscar archivos Excel en la carpeta
        archivos_excel = list(carpeta.glob('*.xlsx')) + list(carpeta.glob('*.xls'))
        
        # Filtrar archivos temporales y archivos consolidados previos
        archivos_excel = [
            f for f in archivos_excel 
            if not f.name.startswith('~$') 
            and not f.name.startswith('Planilla Metso Consolidado')
        ]
        
        if not archivos_excel:
            raise Exception(f"No se encontraron archivos Excel en: {carpeta}")
        
        print(f"✓ Se encontraron {len(archivos_excel)} archivo(s) Excel")
    else:
        raise Exception("Formato de entrada inválido para consolidar_archivos()")
    
    # ====================================================================
    # PROCESAMIENTO DE ARCHIVOS
    # ====================================================================
    
    dataframes = []
    archivos_procesados = 0
    archivos_con_error = []
    
    print(f"\n[1/3] Procesando {len(archivos_excel)} archivo(s)...")
    
    for idx, archivo in enumerate(archivos_excel, 1):
        try:
            # Extraer periodo del nombre del archivo
            periodo = extraer_periodo(archivo.name)
            
            if periodo is None:
                print(f"  [{idx}/{len(archivos_excel)}] ⚠️  {archivo.name} - No se pudo extraer periodo, omitiendo...")
                archivos_con_error.append((archivo.name, "No se pudo extraer periodo"))
                continue
            
            print(f"  [{idx}/{len(archivos_excel)}] Procesando: {archivo.name} (Periodo: {periodo})", end='', flush=True)
            
            # Leer archivo
            df = leer_archivo_planilla(archivo, periodo)
            dataframes.append(df)
            archivos_procesados += 1
            
            print(f" ✓ ({len(df)} registros)")
            
        except Exception as e:
            print(f" ✗ ERROR")
            print(f"      Error: {str(e)}")
            archivos_con_error.append((archivo.name, str(e)))
    
    if not dataframes:
        raise Exception("No se pudo procesar ningún archivo correctamente")
    
    print(f"\n  ✓ Archivos procesados exitosamente: {archivos_procesados}/{len(archivos_excel)}")
    
    if archivos_con_error:
        print(f"  ⚠️  Archivos con error: {len(archivos_con_error)}")
        for nombre, error in archivos_con_error:
            print(f"      - {nombre}: {error}")
    
    # ====================================================================
    # CONSOLIDACIÓN DE DATAFRAMES
    # ====================================================================
    
    print(f"\n[2/3] Consolidando datos...")
    
    # Convertir todas las columnas a string para evitar conflictos de tipos
    print(f"  - Normalizando tipos de datos...")
    dataframes_normalizados = []
    
    for df in dataframes:
        # Convertir todas las columnas (excepto PERIODO) a string
        columnas_a_convertir = [col for col in df.columns if col != "PERIODO"]
        df_normalizado = df.with_columns([
            pl.col(col).cast(pl.Utf8, strict=False).alias(col)
            for col in columnas_a_convertir
        ])
        dataframes_normalizados.append(df_normalizado)
    
    # Usar diagonal=True para manejar columnas diferentes entre archivos
    print(f"  - Concatenando archivos...")
    df_consolidado = pl.concat(dataframes_normalizados, how="diagonal")
    
    print(f"  ✓ Consolidación completa: {len(df_consolidado):,} registros totales")
    print(f"  ✓ Total de columnas: {len(df_consolidado.columns)}")
    
    # ====================================================================
    # LIMPIEZA Y ENRIQUECIMIENTO
    # ====================================================================
    
    print(f"\n  - Limpiando datos...")
    registros_antes = len(df_consolidado)
    
    if "DNI/CEX" in df_consolidado.columns:
        df_consolidado = df_consolidado.filter(pl.col("DNI/CEX").is_not_null())
        registros_despues = len(df_consolidado)
        registros_eliminados = registros_antes - registros_despues
        
        if registros_eliminados > 0:
            print(f"  ✓ Eliminadas {registros_eliminados:,} filas con DNI/CEX nulo")
        else:
            print(f"  ✓ No se encontraron filas con DNI/CEX nulo")
    else:
        print(f"  ⚠️  Advertencia: No se encontró la columna 'DNI/CEX' para limpieza")
    
    # Agregar columnas MES y AÑO derivadas de PERIODO
    print(f"  - Generando columnas MES y AÑO desde PERIODO...")
    
    # Extraer AÑO (primeros 4 caracteres) y MES (últimos 2 caracteres después del guión)
    df_consolidado = df_consolidado.with_columns([
        pl.col("PERIODO").str.slice(0, 4).alias("AÑO"),
        pl.col("PERIODO").str.slice(5, 2).cast(pl.Int32, strict=False).alias("MES")
    ])
    
    # Reorganizar columnas: PERIODO, AÑO, MES, resto
    columnas_ordenadas = ["PERIODO", "AÑO", "MES"] + [
        col for col in df_consolidado.columns if col not in ["PERIODO", "AÑO", "MES"]
    ]
    df_consolidado = df_consolidado.select(columnas_ordenadas)
    
    print(f"  ✓ Columnas MES y AÑO generadas exitosamente")
    
    # ====================================================================
    # GUARDAR RESULTADOS - CORRECCIÓN CRÍTICA
    # ====================================================================
    print(f"\n[3/3] Guardando resultados en capa Silver...")
    
    try:
        # Llamar a la función guardar_resultados para generar Parquet y Excel
        ruta_parquet, ruta_excel = guardar_resultados(df_consolidado, carpeta_trabajo)
        
        # Verificar que los archivos fueron creados
        if not ruta_parquet.exists():
            raise FileNotFoundError(f"No se pudo crear el archivo Parquet: {ruta_parquet}")
        
        if not ruta_excel.exists():
            print(f"  ⚠️  Advertencia: No se pudo crear el archivo Excel: {ruta_excel}")
        else:
            print(f"  ✓ Excel generado: {ruta_excel.name}")
            
    except Exception as e:
        print(f"  ✗ ERROR al guardar resultados: {e}")
        print(f"  [DEBUG] Traceback:")
        traceback.print_exc()
        raise
    
    return df_consolidado


def guardar_resultados(df, carpeta_trabajo):
    """
    Guarda el DataFrame consolidado como parquet y Excel en carpeta silver/
    Sin timestamp - se sobreescribe en cada ejecución
    
    Args:
        df: DataFrame consolidado
        carpeta_trabajo: Path de la carpeta de trabajo
        
    Returns:
        tuple: (ruta_parquet, ruta_excel)
    """
    # Crear carpeta silver si no existe
    carpeta_silver = Path(carpeta_trabajo) / "silver"
    carpeta_silver.mkdir(exist_ok=True)
    
    # Nombres fijos sin timestamp
    nombre_parquet = "Planilla_Metso_Consolidado_Silver.parquet"
    nombre_excel = "Planilla_Metso_Consolidado_Silver.xlsx"
    
    # Rutas de salida
    ruta_parquet = carpeta_silver / nombre_parquet
    ruta_excel = carpeta_silver / nombre_excel
    
    print(f"  📁 Carpeta: {carpeta_silver}")
    
    # Guardar como Parquet (CRÍTICO para el pipeline)
    print(f"  - Guardando parquet...", end='', flush=True)
    try:
        df.write_parquet(ruta_parquet)
        size_mb = ruta_parquet.stat().st_size / (1024 * 1024) if ruta_parquet.exists() else 0
        print(f" ✓ ({size_mb:.2f} MB)")
    except Exception as e:
        print(f" ✗ ERROR: {e}")
        # Intentar con diferentes opciones de compresión
        try:
            print(f"  - Intentando con compresión diferente...")
            df.write_parquet(ruta_parquet, compression='snappy')
            size_mb = ruta_parquet.stat().st_size / (1024 * 1024)
            print(f"  ✓ Parquet guardado con compresión snappy ({size_mb:.2f} MB)")
        except Exception as e2:
            print(f"  ✗ Error también con compresión: {e2}")
            raise
    
    # Guardar como Excel (opcional pero útil para visualización)
    print(f"  - Guardando Excel...", end='', flush=True)
    try:
        df.write_excel(ruta_excel)
        size_mb = ruta_excel.stat().st_size / (1024 * 1024) if ruta_excel.exists() else 0
        print(f" ✓ ({size_mb:.2f} MB)")
    except Exception as e:
        print(f" ✗ ERROR (Excel): {e}")
        # No lanzar excepción para Excel, el pipeline solo necesita Parquet
    
    return ruta_parquet, ruta_excel


# ============================================================================
# FUNCIÓN MAIN PARA EJECUCIÓN STANDALONE
# ============================================================================

def main():
    """Función main para ejecución standalone con selección de carpeta"""
    from tkinter import Tk, filedialog
    
    print("=" * 70)
    print(" CONSOLIDADOR DE PLANILLAS METSO - CAPA SILVER ".center(70, "="))
    print("=" * 70)
    
    # 1. Seleccionar carpeta
    print("\n[PASO 1] Selecciona la carpeta con los archivos de planilla...")
    
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    carpeta = filedialog.askdirectory(title="Selecciona la carpeta con los archivos de planilla")
    root.destroy()
    
    if not carpeta:
        print("✗ No se seleccionó ninguna carpeta. Proceso cancelado.")
        return
    
    # Iniciar cronómetro después de la selección
    tiempo_inicio = time.time()
    
    print(f"✓ Carpeta seleccionada: {carpeta}")
    carpeta_path = Path(carpeta)
    
    # 2. Mostrar archivos encontrados (previsualización)
    print("\n[PASO 2] Buscando archivos Excel...")
    archivos_excel = list(carpeta_path.glob('*.xlsx')) + list(carpeta_path.glob('*.xls'))
    
    # Filtrar archivos temporales y archivos consolidados previos
    archivos_excel = [
        f for f in archivos_excel 
        if not f.name.startswith('~$') 
        and not f.name.startswith('Planilla Metso Consolidado')
    ]
    
    if not archivos_excel:
        print("✗ No se encontraron archivos Excel en la carpeta seleccionada")
        return
    
    print(f"✓ Se encontraron {len(archivos_excel)} archivo(s) Excel")
    
    # Mostrar archivos encontrados
    print("\nArchivos a procesar:")
    for idx, archivo in enumerate(archivos_excel, 1):
        periodo = extraer_periodo(archivo.name)
        if periodo:
            print(f"  {idx}. {archivo.name} → Periodo: {periodo}")
        else:
            print(f"  {idx}. {archivo.name} → ⚠️  No se detectó periodo")
    
    # 3. Consolidar archivos (pasando carpeta en modo standalone)
    print("\n" + "=" * 70)
    print(" PROCESAMIENTO ".center(70, "="))
    print("=" * 70)
    
    try:
        df_consolidado = consolidar_archivos(carpeta_path, carpeta_path)
        
        # NOTA: consolidar_archivos() ya guarda los resultados, no es necesario llamar a guardar_resultados()
        
        # Calcular tiempo total
        tiempo_total = time.time() - tiempo_inicio
        
        # Resumen final
        print("\n" + "=" * 70)
        print(" RESUMEN ".center(70, "="))
        print("=" * 70)
        
        print(f"\n✓ Consolidación completada exitosamente")
        print(f"\n📊 Estadísticas:")
        print(f"  - Total de registros (final): {len(df_consolidado):,}")
        print(f"  - Total de columnas: {len(df_consolidado.columns)}")
        print(f"  - Periodos únicos: {df_consolidado['PERIODO'].n_unique()}")
        
        # Mostrar periodos encontrados
        periodos = df_consolidado['PERIODO'].unique().sort().to_list()
        print(f"  - Periodos: {', '.join(periodos)}")
        
        # Verificar archivos generados
        carpeta_silver = carpeta_path / "silver"
        ruta_parquet = carpeta_silver / "Planilla Metso Consolidado.parquet"
        ruta_excel = carpeta_silver / "Planilla Metso Consolidado.xlsx"
        
        print(f"\n📁 Archivos generados en carpeta silver/:")
        if ruta_parquet.exists():
            size_mb = ruta_parquet.stat().st_size / (1024 * 1024)
            print(f"  ✓ Parquet: {ruta_parquet.name} ({size_mb:.2f} MB)")
        else:
            print(f"  ✗ Parquet: NO GENERADO")
            
        if ruta_excel.exists():
            size_mb = ruta_excel.stat().st_size / (1024 * 1024)
            print(f"  ✓ Excel: {ruta_excel.name} ({size_mb:.2f} MB)")
        else:
            print(f"  ⚠️  Excel: NO GENERADO")
        
        print(f"\n⏱️  Tiempo de ejecución: {tiempo_total:.2f}s")
        
        print("\n💡 Los archivos se sobreescriben en cada ejecución (sin histórico)")
        
        print("\n" + "=" * 70)
        
    except Exception as e:
        print(f"\n✗ Error durante el procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return


if __name__ == "__main__":
    main()