"""
Script: step1_consolidar_ingresos.py
Descripci√≥n: Procesa archivos Excel de Relaci√≥n de Ingresos - Bronze ‚Üí Silver
             - Hoja EMPLEADOS: 22 columnas, headers en fila 2 (empieza en columna B)
             - Hoja PRACTICANTES: 20 columnas, headers en fila 2 (empieza en columna A)
             
Limpieza aplicada:
    - PROYECTO: Reemplazar "0", null o vac√≠o por "Staff"
    - CODIGO SAP: Reemplazar "#N/D" o "Error" por null
    - Genera columna PERIODO desde A√ëO-MES
    
Salida: Archivos sin timestamp en carpeta silver/
    - Relacion Ingresos EMPLEADOS.parquet
    - Relacion Ingresos PRACTICANTES.parquet

Autor: Richi
Fecha: 06.01.2025
"""

import polars as pl
import openpyxl
from pathlib import Path
from datetime import datetime
import sys
from tkinter import Tk, filedialog
import time

# ============================================================================
# CONFIGURACI√ìN
# ============================================================================

CONFIGURACION_HOJAS = {
    "EMPLEADOS": {
        "fila_headers": 2,
        "columna_inicio": 2,  # Columna B (1-indexed)
        "columnas_esperadas": [
            "A√ëO", "MES", "DNI", "N¬∞ DOCUM.", "NOMBRES COMPLETOS", "SEXO",
            "DEPARTAMENTO", "PROVINCIA", "DISTRITO", "CARGO", "√ÅREA", 
            "JEFE DIRECTO", "LUGAR DE TRABAJO", "FECHA INICIO", "CC",
            "WC/BC", "Tiempo de contrato", "NIVEL II", "NIVEL III", 
            "NIVEL IV", "PROYECTO", "CODIGO SAP"
        ]
    },
    "PRACTICANTES": {
        "fila_headers": 2,
        "columna_inicio": 1,  # Columna A (1-indexed)
        "columnas_esperadas": [
            "A√ëO", "MES", "Tipo Documento", "Numero Documento", "CODIGO SAP",
            "Nombres Completos", "Fecha Nacimiento", "Sexo", "Departamento",
            "Provincia", "Distrito", "Nacionalidad", "Cargo", "√Årea",
            "Jefe responsable de √°rea", "Sede de Trabajo", "Fecha Inicio",
            "CC", "Tel√©fono", "Universidad de Procedencia"
        ]
    }
}

# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================

def seleccionar_archivo_excel() -> Path | None:
    """Abre di√°logo para seleccionar archivo Excel"""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    archivo = filedialog.askopenfilename(
        title="Seleccionar archivo de Relaci√≥n de Ingresos (Bronze)",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    
    root.destroy()
    
    return Path(archivo) if archivo else None


def leer_hoja_excel(ruta_excel: Path, nombre_hoja: str, config: dict) -> pl.DataFrame:
    """
    Lee una hoja Excel usando openpyxl con configuraci√≥n espec√≠fica
    
    Args:
        ruta_excel: Ruta al archivo Excel
        nombre_hoja: Nombre de la hoja a leer
        config: Diccionario con fila_headers y columna_inicio
    
    Returns:
        DataFrame de Polars con los datos
    """
    print(f"\nüìñ Leyendo hoja: {nombre_hoja}")
    print(f"   Archivo: {ruta_excel.name}")
    print(f"   Headers en fila: {config['fila_headers']}, columna inicio: {config['columna_inicio']}")
    
    try:
        # Cargar workbook
        wb = openpyxl.load_workbook(ruta_excel, data_only=True, read_only=True)
        
        if nombre_hoja not in wb.sheetnames:
            raise ValueError(f"La hoja '{nombre_hoja}' no existe. Hojas disponibles: {wb.sheetnames}")
        
        ws = wb[nombre_hoja]
        fila_headers = config['fila_headers']
        col_inicio = config['columna_inicio']
        
        # Extraer encabezados desde la columna especificada
        headers = []
        col_idx = col_inicio
        while True:
            cell = ws.cell(row=fila_headers, column=col_idx)
            if cell.value is None:
                break
            headers.append(str(cell.value).strip())
            col_idx += 1
        
        print(f"   ‚úì Encabezados encontrados: {len(headers)} columnas")
        
        # Validar columnas esperadas
        columnas_esperadas = config['columnas_esperadas']
        if len(headers) != len(columnas_esperadas):
            print(f"   ‚ö†Ô∏è  Advertencia: Se esperaban {len(columnas_esperadas)} columnas, se encontraron {len(headers)}")
        
        # Extraer datos
        datos = []
        fila_datos_inicio = fila_headers + 1
        
        for row in ws.iter_rows(min_row=fila_datos_inicio, values_only=True):
            # Tomar datos desde la columna de inicio
            fila_datos = list(row[col_inicio-1:col_inicio-1+len(headers)])
            
            # Detener si encontramos fila completamente vac√≠a
            if all(cell is None or str(cell).strip() == "" for cell in fila_datos):
                break
            
            datos.append(fila_datos)
        
        wb.close()
        
        print(f"   ‚úì Filas de datos extra√≠das: {len(datos)}")
        
        # Crear DataFrame de Polars
        if not datos:
            print("   ‚ö†Ô∏è  No se encontraron datos")
            return pl.DataFrame()
        
        # Convertir todos los valores a string para evitar conflictos de tipo
        datos_str = []
        for fila in datos:
            fila_str = [str(valor) if valor is not None else None for valor in fila]
            datos_str.append(fila_str)
        
        # strict=False permite tipos mixtos en columnas
        df = pl.DataFrame(
            {header: [fila[i] for fila in datos_str] for i, header in enumerate(headers)},
            strict=False
        )
        
        return df
        
    except Exception as e:
        print(f"   ‚ùå Error al leer Excel: {str(e)}")
        raise


def limpiar_datos(df: pl.DataFrame, nombre_hoja: str) -> pl.DataFrame:
    """
    Aplica reglas de limpieza espec√≠ficas para Relaci√≥n de Ingresos
    
    Reglas:
    - PROYECTO: Reemplazar "0" y valores vac√≠os/null por "Staff"
    - CODIGO SAP: Reemplazar "#N/D" o "Error" por null
    - Genera columna PERIODO desde A√ëO-MES
    """
    print(f"\nüßπ Limpiando datos de {nombre_hoja}...")
    
    df_limpio = df
    stats = {"proyecto": 0, "codigo_sap": 0, "filas_vacias": 0}
    
    # Limpieza de PROYECTO (solo en EMPLEADOS)
    if "PROYECTO" in df.columns:
        registros_antes = df_limpio.filter(
            (pl.col("PROYECTO").cast(pl.Utf8) == "0") | 
            (pl.col("PROYECTO").is_null()) |
            (pl.col("PROYECTO").cast(pl.Utf8).str.strip_chars() == "")
        ).height
        
        df_limpio = df_limpio.with_columns(
            pl.when(
                (pl.col("PROYECTO").cast(pl.Utf8) == "0") |
                (pl.col("PROYECTO").is_null()) |
                (pl.col("PROYECTO").cast(pl.Utf8).str.strip_chars() == "")
            )
            .then(pl.lit("Staff"))
            .otherwise(pl.col("PROYECTO"))
            .alias("PROYECTO")
        )
        
        stats["proyecto"] = registros_antes
        print(f"   ‚úì PROYECTO: {registros_antes} valores '0'/null/vac√≠os ‚Üí 'Staff'")
    
    # Limpieza de CODIGO SAP (ambas hojas)
    if "CODIGO SAP" in df.columns:
        registros_antes = df_limpio.filter(
            pl.col("CODIGO SAP").cast(pl.Utf8).is_in(["#N/D", "Error", "#N/A"]) |
            pl.col("CODIGO SAP").cast(pl.Utf8).str.contains("(?i)error") |
            pl.col("CODIGO SAP").is_null()
        ).height
        
        df_limpio = df_limpio.with_columns(
            pl.when(
                (pl.col("CODIGO SAP").cast(pl.Utf8).str.contains("#N/D|#N/A")) |
                (pl.col("CODIGO SAP").cast(pl.Utf8).str.to_lowercase().str.contains("error")) |
                (pl.col("CODIGO SAP").is_null())
            )
            .then(None)
            .otherwise(pl.col("CODIGO SAP"))
            .alias("CODIGO SAP")
        )
        
        stats["codigo_sap"] = registros_antes
        print(f"   ‚úì CODIGO SAP: {registros_antes} valores '#N/D'/'Error' ‚Üí null")
    
    # Limpieza general: eliminar filas completamente vac√≠as
    registros_antes = df_limpio.height
    df_limpio = df_limpio.filter(
        ~pl.all_horizontal(pl.all().is_null())
    )
    stats["filas_vacias"] = registros_antes - df_limpio.height
    
    if stats["filas_vacias"] > 0:
        print(f"   ‚úì Eliminadas {stats['filas_vacias']} filas completamente vac√≠as")
    
    # Generar columna PERIODO (YYYY-MM) desde A√ëO y MES
    if "A√ëO" in df_limpio.columns and "MES" in df_limpio.columns:
        print(f"   - Generando columna PERIODO desde A√ëO-MES...")
        
        df_limpio = df_limpio.with_columns(
            (pl.col("A√ëO").cast(pl.Utf8) + "-" + 
             pl.col("MES").cast(pl.Utf8).str.zfill(2)).alias("PERIODO")
        )
        
        # Reorganizar columnas: PERIODO al inicio
        columnas_ordenadas = ["PERIODO"] + [col for col in df_limpio.columns if col != "PERIODO"]
        df_limpio = df_limpio.select(columnas_ordenadas)
        
        print(f"   ‚úì Columna PERIODO generada exitosamente")
    
    return df_limpio


def generar_reporte_calidad(df_original: pl.DataFrame, df_limpio: pl.DataFrame, nombre_hoja: str):
    """Genera reporte de calidad de datos por hoja"""
    print(f"\nüìä REPORTE DE CALIDAD: {nombre_hoja}")
    print("=" * 80)
    print(f"Registros originales:  {df_original.height:,}")
    print(f"Registros limpios:     {df_limpio.height:,}")
    print(f"Registros eliminados:  {df_original.height - df_limpio.height:,}")
    
    if df_original.height > 0:
        print(f"Tasa de retenci√≥n:     {(df_limpio.height / df_original.height * 100):.2f}%")
    
    print("=" * 80)
    
    # Estad√≠sticas de nulos por columna (solo columnas con nulos)
    print("\nüìã Valores nulos por columna:")
    tiene_nulos = False
    for col in df_limpio.columns:
        nulos = df_limpio[col].is_null().sum()
        if nulos > 0:
            pct = (nulos / df_limpio.height * 100) if df_limpio.height > 0 else 0
            print(f"   {col:35} : {nulos:5,} ({pct:5.2f}%)")
            tiene_nulos = True
    
    if not tiene_nulos:
        print("   ‚úì No hay valores nulos")


def guardar_resultados(resultados: dict, carpeta_trabajo: Path):
    """
    Guarda ambas hojas en silver/ sin timestamp (solo parquet)
    - EMPLEADOS: Se procesar√° despu√©s a Gold
    - PRACTICANTES: Se queda aqu√≠ (solo consulta)
    
    Args:
        resultados: Diccionario con DataFrames por hoja
        carpeta_trabajo: Path de la carpeta de trabajo
        
    Returns:
        dict: Rutas de archivos guardados por hoja
    """
    # Crear carpeta silver si no existe
    carpeta_silver = carpeta_trabajo / "silver"
    carpeta_silver.mkdir(exist_ok=True)
    
    print(f"\n[3/3] Guardando resultados en capa Silver...")
    print(f"  üìÅ Carpeta: {carpeta_silver}")
    
    rutas_guardadas = {}
    
    for nombre_hoja, datos in resultados.items():
        nombre_base = f"Relacion Ingresos {nombre_hoja}"
        
        # Guardar Parquet
        print(f"\n  üìÑ {nombre_hoja}:")
        print(f"    - Guardando parquet...", end='', flush=True)
        ruta_parquet = carpeta_silver / f"{nombre_base}.parquet"
        datos["df"].write_parquet(ruta_parquet, compression="snappy")
        print(f" ‚úì")
        
        print(f"    - Registros: {datos['registros']:,}")
        
        rutas_guardadas[nombre_hoja] = {
            "parquet": ruta_parquet
        }
    
    return rutas_guardadas


# ============================================================================
# FUNCI√ìN PRINCIPAL
# ============================================================================

def main():
    """Funci√≥n principal de procesamiento"""
    print("=" * 80)
    print(" CONSOLIDADOR DE RELACI√ìN DE INGRESOS - CAPA SILVER ".center(80, "="))
    print("=" * 80)
    
    # 1. Seleccionar archivo
    print("\n[PASO 1] Selecciona el archivo de Relaci√≥n de Ingresos (Bronze)...")
    archivo_bronze = seleccionar_archivo_excel()
    
    if not archivo_bronze:
        print("‚ùå No se seleccion√≥ ning√∫n archivo. Proceso cancelado.")
        return
    
    # Iniciar cron√≥metro despu√©s de la selecci√≥n
    tiempo_inicio = time.time()
    
    print(f"‚úì Archivo seleccionado: {archivo_bronze.name}")
    carpeta_trabajo = archivo_bronze.parent
    
    # 2. Procesar hojas
    print("\n" + "=" * 80)
    print(" PROCESAMIENTO ".center(80, "="))
    print("=" * 80)
    print(f"\n[1/3] Procesando hojas...")
    
    resultados = {}
    
    for idx, (nombre_hoja, config) in enumerate(CONFIGURACION_HOJAS.items(), 1):
        print(f"\n{'='*80}")
        print(f"[{idx}/{len(CONFIGURACION_HOJAS)}] PROCESANDO HOJA: {nombre_hoja}")
        print('='*80)
        
        try:
            # 2.1 Leer datos
            df_original = leer_hoja_excel(archivo_bronze, nombre_hoja, config)
            
            if df_original.is_empty():
                print(f"   ‚ö†Ô∏è  No se encontraron datos en {nombre_hoja}")
                continue
            
            # 2.2 Limpiar datos
            df_limpio = limpiar_datos(df_original, nombre_hoja)
            
            # 2.3 Generar reporte de calidad
            generar_reporte_calidad(df_original, df_limpio, nombre_hoja)
            
            # 2.4 Guardar resultados
            resultados[nombre_hoja] = {
                "df": df_limpio,
                "registros": df_limpio.height
            }
            
        except Exception as e:
            print(f"\n‚ùå Error procesando {nombre_hoja}: {str(e)}")
            continue
    
    # 3. Guardar archivos procesados
    if not resultados:
        print("\n‚ùå No se procesaron datos de ninguna hoja.")
        return
    
    print(f"\n{'='*80}")
    print("[2/3] Validando resultados...")
    print('='*80)
    print(f"  ‚úì Hojas procesadas exitosamente: {len(resultados)}/{len(CONFIGURACION_HOJAS)}")
    
    rutas_guardadas = guardar_resultados(resultados, carpeta_trabajo)
    
    # Calcular tiempo total
    tiempo_total = time.time() - tiempo_inicio
    
    # 4. Resumen final
    print("\n" + "=" * 80)
    print(" RESUMEN ".center(80, "="))
    print("=" * 80)
    
    print(f"\n‚úì Consolidaci√≥n completada exitosamente")
    print(f"\nüìä Estad√≠sticas:")
    print(f"  - Hojas procesadas: {len(resultados)}")
    print(f"  - Total de registros: {sum(d['registros'] for d in resultados.values()):,}")
    
    print(f"\nüìÅ Archivos generados en carpeta silver/:")
    for nombre_hoja, rutas in rutas_guardadas.items():
        print(f"\n  {nombre_hoja}:")
        print(f"    - Parquet: {rutas['parquet'].name}")
    
    print(f"\n‚è±Ô∏è  Tiempo de ejecuci√≥n: {tiempo_total:.2f}s")
    
    print("\nüí° Notas:")
    print("  - EMPLEADOS: Listo para procesamiento a Gold (step2)")
    print("  - PRACTICANTES: Se mantiene en Silver (solo consulta)")
    print("  - Archivos sin timestamp (se sobreescriben en cada ejecuci√≥n)")
    
    print("\n" + "=" * 80)


# ============================================================================
# PUNTO DE ENTRADA
# ============================================================================

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n‚ùå Error fatal: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
