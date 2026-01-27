"""
Ejecutor de Pipeline Control de Practicantes
Lee archivo YAML y ejecuta 2 stages secuencialmente:
  1. Bronze → Silver (procesamiento Excel)
  2. Silver → Gold (flags y tiempo de servicio)

Compatible con señales Qt para integración en UI
"""

import yaml
from pathlib import Path
from typing import Dict, Any
from importlib import import_module
from PySide6.QtCore import QObject, Signal
import sys
import time


class PipelineControlPracticantesExecutor(QObject):
    """
    Ejecutor de pipeline basado en YAML para Control de Practicantes
    Emite señales Qt para integración con UILogger
    """
    
    # Señales Qt
    log_message = Signal(str, str)  # (nivel, mensaje)
    progress_update = Signal(int, str)  # (porcentaje, mensaje)
    stage_started = Signal(str, str)  # (nombre_stage, descripción)
    stage_completed = Signal(str, bool, float)  # (nombre_stage, éxito, duración)
    
    def __init__(self, yaml_path: Path, archivo: Path, output_dir: Path):
        """
        Inicializa el executor
        
        Args:
            yaml_path: Ruta al archivo YAML del pipeline
            archivo: Archivo Excel de control de practicantes
            output_dir: Directorio base de salida (carpeta del archivo)
        """
        super().__init__()
        self.yaml_path = yaml_path
        self.archivo = archivo
        self.output_dir = output_dir
        self.pipeline_config = None
        self.stages_results = {}
        
    def _log(self, nivel: str, mensaje: str):
        """Emite señal de log"""
        self.log_message.emit(nivel, mensaje)
    
    def _progress(self, porcentaje: int, mensaje: str):
        """Emite señal de progreso"""
        self.progress_update.emit(porcentaje, mensaje)
    
    def load_pipeline(self) -> bool:
        """
        Carga la configuración del pipeline desde YAML
        
        Returns:
            True si se cargó correctamente, False en caso contrario
        """
        try:
            with open(self.yaml_path, 'r', encoding='utf-8') as f:
                self.pipeline_config = yaml.safe_load(f)
            
            self._log("INFO", f"✓ Pipeline cargado: {self.pipeline_config['name']}")
            self._log("INFO", f"  • Versión: {self.pipeline_config['version']}")
            self._log("INFO", f"  • Stages: {len(self.pipeline_config['stages'])}")
            
            return True
            
        except Exception as e:
            self._log("ERROR", f"Error al cargar pipeline YAML: {e}")
            return False
    
    def validate_structure(self) -> bool:
        """
        Valida que exista el archivo de entrada y tenga la pestaña 'Practicantes'
        
        Returns:
            True si la estructura es válida
        """
        # Validar que el archivo existe
        if not self.archivo.exists():
            self._log("ERROR", f"No se encontró el archivo: {self.archivo.name}")
            return False
        
        self._log("INFO", f"✓ Archivo encontrado: {self.archivo.name}")
        
        # Validar nombre del archivo
        nombre_esperado = "LISTA DE CONTRATOS Y PRACTICANTES - CONTROL"
        if nombre_esperado not in self.archivo.stem:
            self._log("WARNING", f"⚠️  Nombre de archivo no coincide exactamente con: {nombre_esperado}")
            self._log("WARNING", f"   Archivo actual: {self.archivo.stem}")
        
        # Validar que tenga la pestaña "Practicantes"
        try:
            import openpyxl
            wb = openpyxl.load_workbook(self.archivo, read_only=True)
            
            if "Practicantes" not in wb.sheetnames:
                self._log("ERROR", f"El archivo no contiene la pestaña 'Practicantes'")
                self._log("ERROR", f"Pestañas disponibles: {wb.sheetnames}")
                wb.close()
                return False
            
            wb.close()
            self._log("INFO", f"✓ Pestaña 'Practicantes' encontrada")
            
        except Exception as e:
            self._log("ERROR", f"Error al validar archivo Excel: {e}")
            return False
        
        self._log("INFO", f"✓ Estructura validada correctamente")
        
        return True
    
    def execute_stage(self, stage_config: Dict, stage_index: int, total_stages: int) -> bool:
        """
        Ejecuta un stage individual del pipeline
        
        Args:
            stage_config: Configuración del stage desde YAML
            stage_index: Índice del stage (0-based)
            total_stages: Total de stages en el pipeline
            
        Returns:
            True si el stage se ejecutó correctamente
        """
        stage_name = stage_config['name']
        stage_desc = stage_config.get('description', '')
        
        # Emitir señal de inicio
        self.stage_started.emit(stage_name, stage_desc)
        
        self._log("INFO", "")
        self._log("INFO", "=" * 70)
        self._log("INFO", f"STAGE {stage_index + 1}/{total_stages}: {stage_name}")
        if stage_desc:
            self._log("INFO", f"Descripción: {stage_desc}")
        self._log("INFO", "=" * 70)
        
        tiempo_inicio = time.time()
        
        try:
            # Importar módulo
            module_path = stage_config['module']
            function_name = stage_config['function']
            
            self._log("INFO", f"→ Importando: {module_path}")
            module = import_module(module_path)
            
            if not hasattr(module, function_name):
                raise AttributeError(
                    f"Módulo '{module_path}' no tiene función '{function_name}'"
                )
            
            func = getattr(module, function_name)
            self._log("INFO", f"✓ Función encontrada: {function_name}()")
            
            # Preparar parámetros según el stage
            params = self._prepare_stage_params(stage_config, stage_index)
            
            self._log("INFO", f"→ Ejecutando stage...")
            
            # Calcular progreso base
            progress_base = int((stage_index / total_stages) * 100)
            self._progress(progress_base, f"Ejecutando: {stage_name}")
            
            # Ejecutar función
            result = func(**params)
            
            # Guardar resultado para stages posteriores
            self.stages_results[stage_name] = result
            
            # Validar outputs si están definidos
            if 'outputs' in stage_config:
                self._validate_outputs(stage_config['outputs'])
            
            duracion = time.time() - tiempo_inicio
            
            self._log("INFO", f"✓ Stage completado exitosamente")
            self._log("INFO", f"  ⏱️  Duración: {duracion:.2f}s")
            self._log("INFO", "-" * 70)
            
            # Emitir señal de completado
            self.stage_completed.emit(stage_name, True, duracion)
            
            # Actualizar progreso
            progress_end = int(((stage_index + 1) / total_stages) * 100)
            self._progress(progress_end, f"✓ {stage_name}")
            
            return True
            
        except Exception as e:
            duracion = time.time() - tiempo_inicio
            
            self._log("ERROR", f"Error en stage '{stage_name}': {str(e)}")
            
            import traceback
            tb_lines = traceback.format_exc().split('\n')
            for line in tb_lines[:10]:
                if line.strip():
                    self._log("DEBUG", f"  {line}")
            
            self.stage_completed.emit(stage_name, False, duracion)
            
            return False
    
    def _prepare_stage_params(self, stage_config: Dict, stage_index: int) -> Dict[str, Any]:
        """
        Prepara parámetros para cada stage según su posición en el pipeline
        
        Args:
            stage_config: Configuración del stage
            stage_index: Índice del stage
            
        Returns:
            Diccionario con parámetros para la función
        """
        if stage_index == 0:  # Stage 1: Bronze → Silver
            # step1_controlpracticantes.procesar_sin_gui(ruta_archivo, carpeta_salida)
            carpeta_silver = self.output_dir / "silver"
            
            return {
                'ruta_archivo': self.archivo,
                'carpeta_salida': carpeta_silver
            }
        
        elif stage_index == 1:  # Stage 2: Silver → Gold
            # step2_controlpracticantes.procesar_sin_gui(ruta_silver, carpeta_gold)
            ruta_silver = self.output_dir / "silver" / "control_practicantes_silver.parquet"
            carpeta_gold = self.output_dir / "gold"
            
            return {
                'ruta_silver': ruta_silver,
                'carpeta_gold': carpeta_gold
            }
        
        else:
            return {}
    
    def _validate_outputs(self, outputs_config: list):
        """
        Valida que los outputs esperados existan
        
        Args:
            outputs_config: Lista de outputs definidos en YAML
        """
        for output in outputs_config:
            path_template = output['path_template']
            required = output.get('required', True)
            
            # Reemplazar variables
            path_str = path_template.replace('${output_dir}', str(self.output_dir))
            output_path = Path(path_str)
            
            if required and not output_path.exists():
                raise FileNotFoundError(f"Output requerido no generado: {output_path}")
            
            if output_path.exists():
                size_mb = output_path.stat().st_size / (1024 * 1024)
                self._log("INFO", f"  ✓ Output generado: {output_path.name} ({size_mb:.2f} MB)")
    
    def execute(self) -> Dict:
        """
        Ejecuta el pipeline completo
        
        Returns:
            Diccionario con resultado de la ejecución
        """
        tiempo_inicio_total = time.time()
        
        # Cargar pipeline
        if not self.load_pipeline():
            return {
                'success': False,
                'error': 'No se pudo cargar el pipeline YAML'
            }
        
        # Validar estructura
        if not self.validate_structure():
            return {
                'success': False,
                'error': 'Validación de archivo fallida'
            }
        
        # Ejecutar stages
        stages = self.pipeline_config['stages']
        total_stages = len(stages)
        
        self._log("INFO", "")
        self._log("INFO", "=" * 70)
        self._log("INFO", f"INICIANDO PIPELINE: {self.pipeline_config['name']}")
        self._log("INFO", f"Total de stages: {total_stages}")
        self._log("INFO", "=" * 70)
        
        for idx, stage_config in enumerate(stages):
            success = self.execute_stage(stage_config, idx, total_stages)
            
            if not success:
                if self.pipeline_config['config'].get('stop_on_error', True):
                    self._log("ERROR", "Pipeline detenido por error en stage")
                    
                    return {
                        'success': False,
                        'error': f"Error en stage: {stage_config['name']}",
                        'completed_stages': idx,
                        'total_stages': total_stages,
                        'duracion_total': time.time() - tiempo_inicio_total
                    }
        
        # Pipeline completado
        duracion_total = time.time() - tiempo_inicio_total
        
        self._log("INFO", "")
        self._log("INFO", "=" * 70)
        self._log("INFO", "PIPELINE COMPLETADO EXITOSAMENTE")
        self._log("INFO", "=" * 70)
        self._log("INFO", f"⏱️  Tiempo total: {duracion_total:.2f}s")
        self._log("INFO", "=" * 70)
        
        self._progress(100, "✓ Pipeline completado")
        
        return {
            'success': True,
            'completed_stages': total_stages,
            'total_stages': total_stages,
            'duracion_total': duracion_total,
            'stages_results': self.stages_results
        }


def get_resource_path(relative_path: str) -> Path:
    """
    Obtiene ruta absoluta de recursos
    Compatible con PyInstaller
    """
    try:
        base_path = Path(sys._MEIPASS)
    except AttributeError:
        base_path = Path(__file__).parent.parent
    
    return base_path / relative_path