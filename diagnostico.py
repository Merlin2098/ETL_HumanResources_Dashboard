import polars as pl
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import json
from tkinter import Tk, filedialog
import xlsxwriter
from openpyxl import load_workbook

def seleccionar_carpeta():
    """
    Abre un explorador para que el usuario seleccione la carpeta
    """
    root = Tk()
    root.withdraw()  # Ocultar la ventana principal de Tkinter
    root.attributes('-topmost', True)  # Traer al frente
    
    carpeta = filedialog.askdirectory(
        title="Seleccione la carpeta con los archivos Excel a diagnosticar"
    )
    
    root.destroy()
    
    if not carpeta:
        print("❌ No se seleccionó ninguna carpeta. Operación cancelada.")
        return None
    
    return carpeta


def extraer_encabezados_fila_6(archivo_excel):
    """
    Extrae los encabezados desde la fila 6 (índice 5) usando openpyxl
    Esto funciona independientemente del formato de tabla
    """
    try:
        wb = load_workbook(archivo_excel, read_only=True, data_only=True)
        
        if 'Planilla' not in wb.sheetnames:
            return None, f"No se encontró la hoja 'Planilla'"
        
        ws = wb['Planilla']
        
        # Leer la fila 6 (índice 6 en openpyxl porque empieza en 1)
        encabezados = []
        fila_6 = ws[6]
        
        for cell in fila_6:
            valor = cell.value
            if valor is not None:
                # Convertir a string y limpiar espacios
                encabezados.append(str(valor).strip())
            else:
                # Si hay una celda vacía, podemos detener o continuar
                # Detener cuando encontramos la primera celda vacía
                break
        
        wb.close()
        
        if not encabezados:
            return None, "No se encontraron encabezados en la fila 6"
        
        return encabezados, None
        
    except Exception as e:
        return None, str(e)


def diagnosticar_estructura_excel_polars(carpeta_input):
    """
    Diagnostica la estructura de múltiples archivos Excel
    Lee encabezados desde fila 6 usando openpyxl (independiente del formato de tabla)
    """
    carpeta = Path(carpeta_input)
    archivos_excel = list(carpeta.glob("*.xlsx")) + list(carpeta.glob("*.xls"))
    
    if not archivos_excel:
        print("⚠ No se encontraron archivos Excel en la carpeta")
        return None, None, None
    
    resultados = {}
    todos_encabezados = set()
    frecuencia_columnas = defaultdict(int)
    
    print(f"\n{'='*70}")
    print(f"DIAGNÓSTICO DE ESTRUCTURA - {len(archivos_excel)} archivos")
    print(f"Carpeta: {carpeta}")
    print(f"Leyendo encabezados desde fila 6 (A6, B6, C6...)")
    print(f"{'='*70}\n")
    
    # 1. Leer encabezados de cada archivo
    archivos_exitosos = 0
    archivos_con_error = 0
    
    for idx, archivo in enumerate(archivos_excel, 1):
        encabezados, error = extraer_encabezados_fila_6(archivo)
        
        if error:
            archivos_con_error += 1
            print(f"[{idx}/{len(archivos_excel)}] ✗ Error en {archivo.name}: {error}")
            resultados[archivo.name] = {
                'error': error,
                'posicion': idx
            }
        else:
            num_columnas = len(encabezados)
            
            # Guardar información
            resultados[archivo.name] = {
                'encabezados': encabezados,
                'num_columnas': num_columnas,
                'posicion': idx,
                'ruta': str(archivo)
            }
            
            # Actualizar sets globales
            todos_encabezados.update(encabezados)
            for col in encabezados:
                frecuencia_columnas[col] += 1
            
            archivos_exitosos += 1
            print(f"[{idx}/{len(archivos_excel)}] ✓ {archivo.name}: {num_columnas} columnas")
    
    print(f"\n{'='*70}")
    print(f"Resumen: {archivos_exitosos} exitosos | {archivos_con_error} con errores")
    print(f"{'='*70}\n")
    
    if archivos_exitosos == 0:
        print("⚠ No se pudo leer ningún archivo correctamente")
        return None, None, None
    
    # 2. Análisis comparativo
    print(f"{'='*70}")
    print("ANÁLISIS COMPARATIVO")
    print(f"{'='*70}\n")
    
    encabezados_master = sorted(todos_encabezados)
    print(f"📊 Total de encabezados únicos: {len(encabezados_master)}")
    print(f"📁 Archivos procesados correctamente: {archivos_exitosos}\n")
    
    # Encontrar columnas universales (presentes en todos los archivos)
    columnas_universales = [
        col for col, freq in frecuencia_columnas.items() 
        if freq == archivos_exitosos
    ]
    print(f"✓ Columnas presentes en TODOS los archivos: {len(columnas_universales)}")
    
    # Encontrar columnas raras (presentes en pocos archivos)
    columnas_raras = [
        (col, freq) for col, freq in frecuencia_columnas.items() 
        if freq < archivos_exitosos * 0.3  # Menos del 30%
    ]
    if columnas_raras:
        print(f"⚠ Columnas poco comunes (< 30% archivos): {len(columnas_raras)}")
        for col, freq in sorted(columnas_raras, key=lambda x: x[1]):
            print(f"   - '{col}': {freq}/{archivos_exitosos} archivos ({(freq/archivos_exitosos)*100:.1f}%)")
        print()
    
    # 3. Crear DataFrames de análisis con Polars
    
    # DataFrame de resumen por archivo
    data_resumen = []
    for nombre_archivo, info in resultados.items():
        if 'error' in info:
            data_resumen.append({
                'Archivo': nombre_archivo,
                'Estado': 'ERROR',
                'Num_Columnas': None,
                'Columnas_Faltantes': None,
                'Columnas_Extra': None,
                'Error': info['error']
            })
        else:
            encabezados_archivo = set(info['encabezados'])
            faltantes = set(columnas_universales) - encabezados_archivo
            extras = encabezados_archivo - set(columnas_universales)
            
            data_resumen.append({
                'Archivo': nombre_archivo,
                'Estado': 'OK',
                'Num_Columnas': info['num_columnas'],
                'Columnas_Faltantes': len(faltantes),
                'Columnas_Extra': len(extras),
                'Lista_Faltantes': ', '.join(sorted(faltantes)) if faltantes else '-',
                'Lista_Extra': ', '.join(sorted(extras)) if extras else '-'
            })
    
    df_resumen = pl.DataFrame(data_resumen)
    
    # DataFrame de matriz de presencia (archivos como columnas)
    data_matriz = []
    for columna in encabezados_master:
        fila = {
            'Columna': columna,
            'Frecuencia': frecuencia_columnas[columna],
            'Porcentaje': f"{(frecuencia_columnas[columna]/archivos_exitosos)*100:.1f}%"
        }
        
        # Agregar presencia en cada archivo
        for nombre_archivo, info in resultados.items():
            if 'encabezados' in info:
                fila[nombre_archivo] = 'Sí' if columna in info['encabezados'] else 'No'
        
        data_matriz.append(fila)
    
    df_matriz = pl.DataFrame(data_matriz)
    
    # DataFrame de frecuencia de columnas
    df_frecuencia = pl.DataFrame({
        'Columna': list(frecuencia_columnas.keys()),
        'Num_Archivos': list(frecuencia_columnas.values()),
        'Porcentaje': [f"{(v/archivos_exitosos)*100:.1f}%" for v in frecuencia_columnas.values()]
    }).sort('Num_Archivos', descending=True)
    
    # DataFrame de encabezados master
    df_master = pl.DataFrame({
        'Posicion': range(1, len(encabezados_master) + 1),
        'Encabezado': encabezados_master
    })
    
    # DataFrame de columnas universales
    df_universales = pl.DataFrame({
        'Posicion': range(1, len(columnas_universales) + 1),
        'Columnas_Universales': sorted(columnas_universales)
    })
    
    # 4. Generar reporte en consola
    print(f"\n{'='*70}")
    print("ARCHIVOS CON DIFERENCIAS")
    print(f"{'='*70}\n")
    
    archivos_con_diferencias = df_resumen.filter(
        (pl.col('Columnas_Faltantes') > 0) | (pl.col('Columnas_Extra') > 0)
    )
    
    if len(archivos_con_diferencias) > 0:
        for row in archivos_con_diferencias.iter_rows(named=True):
            print(f"📄 {row['Archivo']}")
            if row['Columnas_Faltantes'] > 0:
                print(f"   ⚠ Faltan {row['Columnas_Faltantes']} columnas: {row['Lista_Faltantes']}")
            if row['Columnas_Extra'] > 0:
                print(f"   ⚠ Tiene {row['Columnas_Extra']} columnas extra: {row['Lista_Extra']}")
            print()
    else:
        print("✓ Todos los archivos tienen la misma estructura\n")
    
    # 5. Exportar a Excel usando xlsxwriter directamente
    timestamp = datetime.now().strftime("%d.%m.%Y_%H.%M.%S")
    output_file = carpeta / f"Diagnostico_Estructura_{timestamp}.xlsx"
    
    print(f"{'='*70}")
    print("EXPORTANDO RESULTADOS...")
    print(f"{'='*70}\n")
    
    # Crear workbook con xlsxwriter
    workbook = xlsxwriter.Workbook(output_file)
    
    # Formatos
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4472C4',
        'font_color': 'white',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })
    
    cell_format = workbook.add_format({
        'border': 1,
        'valign': 'vcenter'
    })
    
    # Función auxiliar para escribir DataFrames
    def escribir_dataframe(worksheet, df, sheet_name):
        """Escribe un DataFrame de Polars a una hoja de Excel"""
        # Escribir encabezados
        for col_num, col_name in enumerate(df.columns):
            worksheet.write(0, col_num, col_name, header_format)
        
        # Escribir datos
        for row_num, row in enumerate(df.iter_rows(), start=1):
            for col_num, value in enumerate(row):
                worksheet.write(row_num, col_num, value, cell_format)
        
        # Autoajustar columnas
        for col_num, col_name in enumerate(df.columns):
            # Calcular ancho basado en el contenido
            max_len = len(str(col_name))
            for val in df[col_name]:
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            worksheet.set_column(col_num, col_num, min(max_len + 2, 60))
    
    # Hoja 1: Resumen
    ws_resumen = workbook.add_worksheet('1_Resumen')
    escribir_dataframe(ws_resumen, df_resumen, '1_Resumen')
    
    # Hoja 2: Frecuencia de columnas
    ws_frecuencia = workbook.add_worksheet('2_Frecuencia_Columnas')
    escribir_dataframe(ws_frecuencia, df_frecuencia, '2_Frecuencia_Columnas')
    
    # Hoja 3: Columnas universales
    ws_universales = workbook.add_worksheet('3_Columnas_Universales')
    escribir_dataframe(ws_universales, df_universales, '3_Columnas_Universales')
    
    # Hoja 4: Matriz de presencia (limitada si hay muchos archivos)
    if len(archivos_excel) <= 50:
        ws_matriz = workbook.add_worksheet('4_Matriz_Presencia')
        escribir_dataframe(ws_matriz, df_matriz, '4_Matriz_Presencia')
    else:
        print(f"⚠ Matriz de presencia omitida (demasiados archivos: {len(archivos_excel)})")
    
    # Hoja 5: Lista maestra
    ws_master = workbook.add_worksheet('5_Encabezados_Master')
    escribir_dataframe(ws_master, df_master, '5_Encabezados_Master')
    
    workbook.close()
    
    print(f"✓ Diagnóstico exportado: {output_file.name}")
    print(f"✓ Hojas creadas: Resumen, Frecuencia, Universales, {'Matriz, ' if len(archivos_excel) <= 50 else ''}Master\n")
    
    # 6. Exportar JSON adicional con metadata completa
    output_json = carpeta / f"Diagnostico_Metadata_{timestamp}.json"
    metadata = {
        'total_archivos': len(archivos_excel),
        'archivos_exitosos': archivos_exitosos,
        'archivos_con_error': archivos_con_error,
        'total_encabezados_unicos': len(encabezados_master),
        'columnas_universales': columnas_universales,
        'timestamp': timestamp,
        'carpeta': str(carpeta),
        'archivos': resultados
    }
    
    with open(output_json, 'w', encoding='utf-8') as f:
        json.dump(metadata, f, indent=2, ensure_ascii=False)
    
    print(f"✓ Metadata exportada: {output_json.name}\n")
    
    return resultados, encabezados_master, df_resumen


def main():
    """
    Función principal que ejecuta el diagnóstico completo
    """
    print("\n" + "="*70)
    print("DIAGNÓSTICO DE ESTRUCTURA DE ARCHIVOS EXCEL")
    print("="*70)
    
    # Seleccionar carpeta
    carpeta = seleccionar_carpeta()
    
    if not carpeta:
        return
    
    # Ejecutar diagnóstico
    inicio = datetime.now()
    resultados, encabezados_master, df_resumen = diagnosticar_estructura_excel_polars(carpeta)
    fin = datetime.now()
    
    # Resumen final
    if resultados:
        print("\n" + "="*70)
        print("PROCESO COMPLETADO")
        print("="*70)
        print(f"⏱ Tiempo de ejecución: {(fin - inicio).total_seconds():.2f} segundos")
        print(f"📊 Archivos procesados: {len(resultados)}")
        print(f"📁 Resultados guardados en: {carpeta}")
        print("="*70 + "\n")
    else:
        print("\n⚠ No se generaron resultados\n")


if __name__ == "__main__":
    main()