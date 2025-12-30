"""
Script de validación de esquema de reportes
Valida encabezados en fila 6 y genera reporte en Excel
"""

import polars as pl
import json
from pathlib import Path
from tkinter import Tk, filedialog
from datetime import datetime
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def seleccionar_carpeta():
    """Abre diálogo para seleccionar carpeta"""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    carpeta = filedialog.askdirectory(title="Selecciona la carpeta con los archivos")
    root.destroy()
    return carpeta


def seleccionar_esquema():
    """Abre diálogo para seleccionar archivo JSON del esquema"""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    archivo = filedialog.askopenfilename(
        title="Selecciona el archivo de esquema JSON",
        filetypes=[("Archivos JSON", "*.json"), ("Todos los archivos", "*.*")]
    )
    root.destroy()
    return archivo


def cargar_esquema(ruta_esquema):
    """Carga el esquema JSON"""
    with open(ruta_esquema, 'r', encoding='utf-8') as f:
        esquema = json.load(f)
    return esquema


def leer_encabezados_fila6(archivo_path):
    """
    Lee los encabezados desde la fila 6 usando openpyxl
    
    Returns:
        list: Lista de encabezados encontrados
    """
    try:
        wb = load_workbook(archivo_path, data_only=True, read_only=True)
        ws = wb.active
        
        # Leer la fila 6 (índice 6)
        encabezados = []
        fila_encabezados = 6
        
        for col in range(1, 100):  # Leer hasta columna 100
            cell = ws.cell(row=fila_encabezados, column=col)
            valor = cell.value
            
            if valor is not None:
                # Limpiar el encabezado
                encabezado = str(valor).strip()
                encabezados.append(encabezado)
            else:
                # Si encontramos 3 columnas vacías consecutivas, asumimos que terminaron los encabezados
                if col > len(encabezados) + 3:
                    break
        
        wb.close()
        return encabezados
        
    except Exception as e:
        raise Exception(f"Error al leer encabezados: {str(e)}")


def analizar_archivo(archivo_path, esquema):
    """
    Analiza un archivo y valida sus encabezados
    
    Returns:
        dict con metadata del archivo
    """
    try:
        # Leer encabezados desde la fila 6
        encabezados_encontrados = leer_encabezados_fila6(archivo_path)
        
        # Validar que la fila 6 tenga datos
        if not encabezados_encontrados or len(encabezados_encontrados) == 0:
            raise Exception("No se encontraron encabezados en la fila 6")
        
        encabezados_esperados = set(esquema['tipos_datos'].keys())
        
        # Calcular estadísticas
        encontrados_set = set(encabezados_encontrados)
        presentes = encontrados_set & encabezados_esperados
        faltantes = encabezados_esperados - encontrados_set
        extras = encontrados_set - encabezados_esperados
        
        # Validar si está en fila 6
        fila_correcta = True
        
        metadata = {
            'nombre_archivo': archivo_path.name,
            'ruta_completa': str(archivo_path),
            'fecha_analisis': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'encabezados_en_fila_6': fila_correcta,
            'total_encabezados_esperados': len(encabezados_esperados),
            'total_encabezados_encontrados': len(encabezados_encontrados),
            'encabezados_presentes': len(presentes),
            'encabezados_faltantes': len(faltantes),
            'encabezados_extras': len(extras),
            'porcentaje_coincidencia': round((len(presentes) / len(encabezados_esperados)) * 100, 2) if len(encabezados_esperados) > 0 else 0.0,
            'validacion_exitosa': len(faltantes) == 0,
            'lista_faltantes': ', '.join(sorted(faltantes)) if faltantes else 'Ninguno',
            'lista_extras': ', '.join(sorted(extras)) if extras else 'Ninguno',
            'lista_presentes': ', '.join(sorted(presentes)) if presentes else 'Ninguno'
        }
        
        return metadata
        
    except Exception as e:
        return {
            'nombre_archivo': archivo_path.name,
            'ruta_completa': str(archivo_path),
            'fecha_analisis': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'encabezados_en_fila_6': False,
            'total_encabezados_esperados': len(esquema['tipos_datos']),
            'total_encabezados_encontrados': 0,
            'encabezados_presentes': 0,
            'encabezados_faltantes': len(esquema['tipos_datos']),
            'encabezados_extras': 0,
            'porcentaje_coincidencia': 0.0,
            'validacion_exitosa': False,
            'lista_faltantes': 'ERROR AL LEER ARCHIVO',
            'lista_extras': '',
            'lista_presentes': '',
            'error': str(e)
        }


def generar_reporte_excel(df_metadata, ruta_salida='reporte_validacion.xlsx'):
    """
    Genera reporte en Excel con formato usando openpyxl
    """
    # Convertir polars DataFrame a pandas para compatibilidad con openpyxl
    df_pandas = df_metadata.to_pandas()
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    validacion_ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    validacion_ok_font = Font(color="006100")
    
    validacion_error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    validacion_error_font = Font(color="9C0006")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escribir encabezados
    for col_idx, column_name in enumerate(df_pandas.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Escribir datos
    for row_idx, row in enumerate(df_pandas.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Aplicar formato condicional a validacion_exitosa
            if df_pandas.columns[col_idx - 1] == 'validacion_exitosa':
                if value:
                    cell.fill = validacion_ok_fill
                    cell.font = validacion_ok_font
                else:
                    cell.fill = validacion_error_fill
                    cell.font = validacion_error_font
            
            # Formato a porcentaje
            if df_pandas.columns[col_idx - 1] == 'porcentaje_coincidencia':
                cell.number_format = '0.00"%"'
                cell.alignment = Alignment(horizontal="center")
    
    # Ajustar anchos de columna
    column_widths = {
        'nombre_archivo': 35,
        'ruta_completa': 55,
        'fecha_analisis': 20,
        'encabezados_en_fila_6': 20,
        'total_encabezados_esperados': 20,
        'total_encabezados_encontrados': 22,
        'encabezados_presentes': 20,
        'encabezados_faltantes': 20,
        'encabezados_extras': 18,
        'porcentaje_coincidencia': 20,
        'validacion_exitosa': 18,
        'lista_faltantes': 45,
        'lista_extras': 45,
        'lista_presentes': 45
    }
    
    for col_idx, column_name in enumerate(df_pandas.columns, 1):
        width = column_widths.get(column_name, 15)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    
    # Guardar
    wb.save(ruta_salida)
    print(f"✓ Reporte generado: {ruta_salida}")


def main():
    print("=" * 70)
    print(" VALIDADOR DE ESQUEMA DE REPORTES ".center(70, "="))
    print("=" * 70)
    
    # Iniciar cronómetro
    tiempo_inicio = time.time()
    
    # 1. Seleccionar archivo de esquema JSON
    print("\n[1/5] Selecciona el archivo de esquema JSON...")
    ruta_esquema = seleccionar_esquema()
    
    if not ruta_esquema:
        print("✗ No se seleccionó ningún archivo de esquema. Proceso cancelado.")
        return
    
    print(f"✓ Esquema seleccionado: {Path(ruta_esquema).name}")
    
    # 2. Cargar esquema
    print("\n[2/5] Cargando esquema...")
    try:
        esquema = cargar_esquema(ruta_esquema)
        print(f"✓ Esquema cargado: {len(esquema['tipos_datos'])} encabezados esperados")
        print(f"  - Versión: {esquema['metadata']['version']}")
        print(f"  - Última actualización: {esquema['metadata']['fecha_actualizacion']}")
        
        # Mostrar algunos encabezados de ejemplo
        print(f"  - Ejemplos: {list(esquema['tipos_datos'].keys())[:3]}...")
        
    except FileNotFoundError:
        print(f"✗ Error: No se encontró el archivo '{ruta_esquema}'")
        return
    except json.JSONDecodeError:
        print(f"✗ Error: El archivo '{ruta_esquema}' no es un JSON válido")
        return
    except KeyError as e:
        print(f"✗ Error: El esquema no tiene la estructura esperada. Falta: {e}")
        return
    
    # 3. Seleccionar carpeta
    print("\n[3/5] Selecciona la carpeta con los archivos...")
    carpeta = seleccionar_carpeta()
    
    if not carpeta:
        print("✗ No se seleccionó ninguna carpeta. Proceso cancelado.")
        return
    
    print(f"✓ Carpeta seleccionada: {carpeta}")
    tiempo_seleccion = time.time()
    
    # 4. Buscar archivos Excel
    carpeta_path = Path(carpeta)
    archivos_excel = list(carpeta_path.glob('*.xlsx')) + list(carpeta_path.glob('*.xls'))
    
    # Filtrar archivos temporales de Excel
    archivos_excel = [f for f in archivos_excel if not f.name.startswith('~$')]
    
    if not archivos_excel:
        print("✗ No se encontraron archivos Excel en la carpeta seleccionada")
        return
    
    print(f"\n[4/5] Analizando {len(archivos_excel)} archivo(s)...")
    
    # 5. Analizar cada archivo
    resultados = []
    for idx, archivo in enumerate(archivos_excel, 1):
        print(f"  [{idx}/{len(archivos_excel)}] Procesando: {archivo.name}", end='', flush=True)
        metadata = analizar_archivo(archivo, esquema)
        resultados.append(metadata)
        
        # Indicador visual
        if metadata['validacion_exitosa']:
            print(" ✓")
        else:
            if 'error' in metadata:
                print(f" ✗ (ERROR: {metadata.get('error', 'Desconocido')})")
            else:
                print(f" ✗ ({metadata['porcentaje_coincidencia']}% coincidencia)")
    
    # 6. Crear DataFrame de Polars
    df_metadata = pl.DataFrame(resultados)
    
    # 7. Guardar como Parquet
    print("\n[5/5] Generando reportes...")
    parquet_path = 'metadata_validacion.parquet'
    df_metadata.write_parquet(parquet_path)
    print(f"✓ Parquet generado: {parquet_path}")
    
    # 8. Exportar a Excel
    excel_path = 'reporte_validacion.xlsx'
    generar_reporte_excel(df_metadata, excel_path)
    
    # Resumen final
    tiempo_total = time.time() - tiempo_inicio
    tiempo_procesamiento = tiempo_total - (tiempo_seleccion - tiempo_inicio)
    
    print("\n" + "=" * 70)
    print(" RESUMEN ".center(70, "="))
    print("=" * 70)
    
    total_archivos = len(df_metadata)
    exitosos = df_metadata.filter(pl.col('validacion_exitosa') == True).height
    fallidos = total_archivos - exitosos
    
    print(f"\nArchivos analizados: {total_archivos}")
    print(f"  ✓ Validaciones exitosas: {exitosos}")
    print(f"  ✗ Validaciones fallidas: {fallidos}")
    
    if total_archivos > 0:
        promedio_coincidencia = df_metadata['porcentaje_coincidencia'].mean()
        print(f"\nPromedio de coincidencia: {promedio_coincidencia:.2f}%")
    
    print(f"\n⏱️  TIEMPO DE EJECUCIÓN:")
    print(f"  - Tiempo de selección: {tiempo_seleccion - tiempo_inicio:.2f}s")
    print(f"  - Tiempo de procesamiento: {tiempo_procesamiento:.2f}s")
    print(f"  - Tiempo total: {tiempo_total:.2f}s")
    print("\n" + "=" * 70)


if __name__ == "__main__":
    main()