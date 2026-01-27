"""
Script: step1_controlpracticantes.py
Descripción: Procesa archivo Excel LISTA_DE_CONTRATOS_Y_PRACTICANTES_-_CONTROL.xlsx
             - Hoja: Practicantes
             - Headers en fila 4 (índice 3)
             - Headers combinados que necesitan limpieza
             
Arquitectura:
- Entrada: Excel con información de practicantes (ubicación seleccionada por usuario)
- Silver: Parquet limpio con columnas estandarizadas

Salida: 
    - /silver/control_practicantes_silver.parquet
    - /silver/control_practicantes_silver.xlsx

Autor: Richi via Claude
Fecha: 27.01.2026
"""

import polars as pl
import openpyxl
from pathlib import Path
from datetime import datetime
import time
import sys
import json
from tkinter import Tk, filedialog


def get_resource_path(relative_path: str) -> Path:
    """
    Obtiene la ruta absoluta de un recurso, manejando tanto
    ejecución standalone como PyInstaller.
    """
    try:
        # PyInstaller crea una carpeta temporal _MEIPASS
        base_path = Path(sys._MEIPASS)
    except AttributeError:
        # Ejecución normal desde el directorio del script
        base_path = Path(__file__).parent.parent

    return base_path / relative_path


def seleccionar_archivo_excel() -> Path | None:
    """Abre diálogo para seleccionar archivo Excel"""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    archivo = filedialog.askopenfilename(
        title="Seleccionar archivo LISTA_DE_CONTRATOS_Y_PRACTICANTES_-_CONTROL.xlsx",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")]
    )
    
    root.destroy()
    
    return Path(archivo) if archivo else None


def cargar_esquema(ruta_esquema: Path) -> dict:
    """Carga el esquema JSON de validación."""
    try:
        with open(ruta_esquema, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"   ✗ Error al cargar esquema: {e}")
        raise


def validar_esquema(df: pl.DataFrame, esquema: dict) -> tuple[bool, list[str]]:
    """
    Valida que el DataFrame cumpla con el esquema definido.
    Retorna (es_valido, lista_errores)
    """
    errores = []
    columnas_df = set(df.columns)
    
    # Validar columnas obligatorias
    columnas_requeridas = esquema.get("required_columns", [])
    
    for col_nombre in columnas_requeridas:
        if col_nombre not in columnas_df:
            errores.append(f"Columna obligatoria faltante: {col_nombre}")
    
    # Validar tipos de dato
    tipos_esperados = esquema.get("column_types", {})
    
    for col_nombre, tipo_esperado in tipos_esperados.items():
        if col_nombre not in columnas_df:
            continue
            
        tipo_actual = str(df[col_nombre].dtype)
        
        # Mapeo flexible de tipos
        if tipo_esperado == "string" and "Utf8" not in tipo_actual and "String" not in tipo_actual:
            errores.append(
                f"Columna '{col_nombre}': tipo esperado string, encontrado {tipo_actual}"
            )
        elif tipo_esperado == "date" and "Date" not in tipo_actual:
            errores.append(
                f"Columna '{col_nombre}': tipo esperado date, encontrado {tipo_actual}"
            )
    
    return len(errores) == 0, errores


def limpiar_nombre_columna(nombre: str) -> str:
    """
    Limpia nombres de columna eliminando espacios excesivos,
    saltos de línea y caracteres especiales.
    """
    if nombre is None:
        return ""
    
    # Convertir a string y limpiar
    nombre_limpio = str(nombre).strip()
    
    # Reemplazar saltos de línea y múltiples espacios
    nombre_limpio = nombre_limpio.replace('\n', ' ').replace('\r', ' ')
    nombre_limpio = ' '.join(nombre_limpio.split())
    
    return nombre_limpio


def leer_hoja_practicantes(
    ruta_archivo: Path,
    esquema: dict
) -> pl.DataFrame | None:
    """
    Lee la hoja 'Practicantes' del archivo Excel.
    Headers en fila 4 (índice 3), datos desde fila 5.
    
    Args:
        ruta_archivo: Path al archivo Excel
        esquema: Diccionario con el esquema de validación
        
    Returns:
        DataFrame con los datos procesados o None si hay error
    """
    print(f"\n   → Procesando hoja: Practicantes")
    
    try:
        wb = openpyxl.load_workbook(ruta_archivo, read_only=True, data_only=True)
        
        if "Practicantes" not in wb.sheetnames:
            print(f"   ✗ Hoja 'Practicantes' no encontrada")
            wb.close()
            return None
        
        ws = wb["Practicantes"]
        
        # Leer headers desde fila 4 (índice 3)
        headers_raw = [cell.value for cell in ws[4]]
        
        # Limpiar headers
        headers = [limpiar_nombre_columna(h) for h in headers_raw if h is not None]
        
        if not headers:
            print(f"   ✗ No se encontraron headers válidos")
            wb.close()
            return None
        
        print(f"   ✓ Headers encontrados: {len(headers)}")
        
        # Columnas objetivo según esquema
        columnas_objetivo = esquema.get("required_columns", [])
        
        # Mapear índices de columnas objetivo
        indices_objetivo = {}
        for idx, header in enumerate(headers_raw):
            if header is not None:
                header_limpio = limpiar_nombre_columna(header)
                if header_limpio in columnas_objetivo:
                    indices_objetivo[header_limpio] = idx
        
        # Verificar que todas las columnas objetivo estén presentes
        columnas_faltantes = set(columnas_objetivo) - set(indices_objetivo.keys())
        if columnas_faltantes:
            print(f"   ✗ Columnas faltantes: {columnas_faltantes}")
            wb.close()
            return None
        
        print(f"   ✓ Todas las columnas objetivo encontradas")
        
        # Leer datos desde fila 5 en adelante
        datos = []
        filas_procesadas = 0
        filas_filtradas = 0
        
        for fila in ws.iter_rows(min_row=5, values_only=True):
            # FILTRO DE SEGURIDAD: DNI no puede ser nulo/vacío
            idx_dni = indices_objetivo.get("DNI", -1)
            
            if idx_dni == -1 or idx_dni >= len(fila):
                filas_filtradas += 1
                continue
                
            dni_valor = fila[idx_dni]
            
            if dni_valor is None or str(dni_valor).strip() == "":
                filas_filtradas += 1
                continue
            
            # Extraer solo columnas objetivo en el orden definido
            fila_filtrada = []
            for col_nombre in columnas_objetivo:
                idx = indices_objetivo.get(col_nombre)
                valor = fila[idx] if idx is not None and idx < len(fila) else None
                fila_filtrada.append(valor)
            
            datos.append(fila_filtrada)
            filas_procesadas += 1
        
        wb.close()
        
        if filas_filtradas > 0:
            print(f"   ⚠️  Filas filtradas por DNI vacío: {filas_filtradas}")
        
        if not datos:
            print(f"   ✗ No se encontraron datos válidos")
            return None
        
        # Crear DataFrame con columnas objetivo
        df = pl.DataFrame(datos, schema=columnas_objetivo, orient="row")
        
        print(f"   ✓ Filas leídas: {filas_procesadas}")
        
        # TRANSFORMACIÓN DE FECHAS: Convertir FECHA ING a formato AAAA-MM-DD
        if "FECHA ING" in df.columns:
            print(f"   → Convirtiendo FECHA ING a formato estándar...")
            
            df = df.with_columns([
                pl.col("FECHA ING")
                  .map_elements(lambda x: convertir_fecha_excel(x), return_dtype=pl.Utf8)
                  .alias("FECHA ING")
            ])
            
            # Convertir a tipo Date
            try:
                df = df.with_columns([
                    pl.col("FECHA ING").str.to_date("%Y-%m-%d", strict=False)
                ])
                print(f"   ✓ FECHA ING convertida a formato Date")
            except Exception as e:
                print(f"   ⚠️  No se pudo convertir FECHA ING a Date: {e}")
        
        # Asegurar que todas las columnas sean string excepto FECHA ING
        for col in df.columns:
            if col != "FECHA ING" and df[col].dtype != pl.Utf8:
                df = df.with_columns([
                    pl.col(col).cast(pl.Utf8, strict=False)
                ])
        
        return df
        
    except Exception as e:
        print(f"   ✗ Error al leer archivo: {e}")
        import traceback
        traceback.print_exc()
        return None


def convertir_fecha_excel(valor) -> str | None:
    """
    Convierte valores de fecha de Excel a formato AAAA-MM-DD.
    Maneja tanto números seriales de Excel como strings de fecha.
    """
    if valor is None:
        return None
    
    # Si ya es un datetime de Python
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    
    # Si es un número (serial date de Excel)
    if isinstance(valor, (int, float)):
        try:
            # Excel serial date: 1 = 1900-01-01
            # Pero Excel tiene un bug con 1900 siendo año bisiesto
            base_date = datetime(1899, 12, 30)
            fecha = base_date + pl.duration(days=int(valor))
            return fecha.strftime("%Y-%m-%d")
        except:
            return None
    
    # Si es string, intentar parsear
    if isinstance(valor, str):
        valor_limpio = valor.strip()
        
        # Intentar varios formatos comunes
        formatos = [
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y/%m/%d",
            "%d.%m.%Y"
        ]
        
        for fmt in formatos:
            try:
                fecha = datetime.strptime(valor_limpio, fmt)
                return fecha.strftime("%Y-%m-%d")
            except:
                continue
    
    return None


def crear_estructura_carpetas(carpeta_trabajo: Path) -> tuple[Path, Path]:
    """
    Crea la estructura de carpetas silver/gold si no existe.
    
    Returns:
        Tuple con (carpeta_silver, carpeta_gold)
    """
    carpeta_silver = carpeta_trabajo / "silver"
    carpeta_gold = carpeta_trabajo / "gold"
    
    carpeta_silver.mkdir(exist_ok=True)
    carpeta_gold.mkdir(exist_ok=True)
    
    return carpeta_silver, carpeta_gold


def exportar_silver(
    df: pl.DataFrame,
    carpeta_silver: Path
):
    """
    Exporta DataFrame a archivos Parquet y Excel en capa Silver.
    
    Args:
        df: DataFrame a exportar
        carpeta_silver: Ruta a la carpeta silver/
    """
    print(f"\n[2/2] Exportando a capa Silver...")
    
    # Exportar Parquet
    ruta_parquet = carpeta_silver / "control_practicantes_silver.parquet"
    df.write_parquet(ruta_parquet, compression="snappy")
    print(f" ✓ Parquet: {ruta_parquet.name}")
    
    # Exportar Excel para visualización
    ruta_excel = carpeta_silver / "control_practicantes_silver.xlsx"
    df.write_excel(ruta_excel)
    print(f" ✓ Excel: {ruta_excel.name}")
    
    print(f"\n📊 Estadísticas:")
    print(f"  - Total registros: {len(df):,}")
    print(f"  - Columnas: {len(df.columns)}")


def main():
    """Función principal de procesamiento"""
    print("=" * 80)
    print(" CONTROL DE PRACTICANTES - BRONZE → SILVER ".center(80, "="))
    print("=" * 80)
    
    # 1. Seleccionar archivo
    print("\n[PASO 1] Selecciona el archivo de control de practicantes...")
    archivo_bronze = seleccionar_archivo_excel()
    
    if not archivo_bronze:
        print("✗ No se seleccionó ningún archivo. Proceso cancelado.")
        return
    
    # Iniciar cronómetro después de la selección
    tiempo_inicio = time.time()
    
    print(f"✓ Archivo seleccionado: {archivo_bronze.name}")
    carpeta_trabajo = archivo_bronze.parent
    
    # 2. Crear estructura de carpetas
    print("\n" + "=" * 80)
    print(" PREPARACIÓN ".center(80, "="))
    print("=" * 80)
    print(f"\n   Creando estructura de carpetas...")
    
    carpeta_silver, carpeta_gold = crear_estructura_carpetas(carpeta_trabajo)
    print(f"   ✓ Carpetas creadas: silver/ gold/")
    
    # 3. Cargar esquema
    print(f"\n[1/2] Cargando esquema de validación...")
    
    try:
        ruta_esquema = get_resource_path("esquemas/esquema_control_practicantes.json")
        esquema = cargar_esquema(ruta_esquema)
        print(f"   ✓ Esquema cargado: {ruta_esquema.name}")
    except Exception as e:
        print(f"✗ Error al cargar esquema: {e}")
        return
    
    # 4. Procesar hoja
    print(f"\n" + "=" * 80)
    print(" PROCESAMIENTO ".center(80, "="))
    print("=" * 80)
    print(f"\n   Extrayendo datos de hoja 'Practicantes'...")
    
    try:
        df = leer_hoja_practicantes(archivo_bronze, esquema)
        
        if df is None:
            print("\n✗ No se pudo procesar la hoja de practicantes")
            return
        
        print(f"\n   ✓ Total registros extraídos: {len(df):,}")
        
        # Validar esquema
        print(f"\n   Validando esquema...")
        es_valido, errores = validar_esquema(df, esquema)
        
        if not es_valido:
            print("   ⚠️  Advertencias de validación:")
            for error in errores:
                print(f"     • {error}")
            # No detener el proceso, solo advertir
        else:
            print(f"   ✓ Esquema válido")
        
        # Exportar resultados
        exportar_silver(df, carpeta_silver)
        
        # Calcular tiempo total
        tiempo_total = time.time() - tiempo_inicio
        
        # 5. Resumen final
        print("\n" + "=" * 80)
        print(" RESUMEN ".center(80, "="))
        print("=" * 80)
        
        print(f"\n✓ Procesamiento completado exitosamente")
        
        print(f"\n📂 Archivos generados en /silver/:")
        print(f"  - control_practicantes_silver.parquet")
        print(f"  - control_practicantes_silver.xlsx")
        
        print(f"\n⏱️  Tiempo de ejecución: {tiempo_total:.2f}s")
        
        print("\n💡 Los archivos se sobreescriben en cada ejecución")
        
        print("\n" + "=" * 80)
        
    except Exception as e:
        print(f"\n✗ Error durante el procesamiento: {str(e)}")
        import traceback
        traceback.print_exc()
        return


def procesar_sin_gui(ruta_archivo: Path, carpeta_salida: Path) -> dict:
    """
    Procesa control de practicantes sin interfaz gráfica (modo headless)
    Compatible con pipeline executor
    
    Args:
        ruta_archivo: Path al archivo Excel de control
        carpeta_salida: Path a carpeta /silver/ donde guardar los resultados
        
    Returns:
        dict con resultados del procesamiento
    """
    print(f"\n🔄 Procesando control de practicantes (modo headless)...")
    print(f"   Archivo: {ruta_archivo.name}")
    print(f"   Salida: {carpeta_salida}")
    
    try:
        # Cargar esquema
        ruta_esquema = get_resource_path("esquemas/esquema_control_practicantes.json")
        esquema = cargar_esquema(ruta_esquema)
        print(f"   ✓ Esquema cargado: {ruta_esquema.name}")
        
        # Procesar hoja Practicantes
        df = leer_hoja_practicantes(ruta_archivo, esquema)
        
        if df is None:
            raise ValueError("No se pudo procesar la hoja 'Practicantes'")
        
        registros_procesados = len(df)
        print(f"   ✓ Registros procesados: {registros_procesados:,}")
        
        # Validar esquema
        es_valido, errores = validar_esquema(df, esquema)
        
        if not es_valido:
            print("   ⚠️  Advertencias de validación:")
            for error in errores:
                print(f"     • {error}")
        
        # Crear carpeta de salida si no existe
        carpeta_salida.mkdir(parents=True, exist_ok=True)
        
        # Guardar Parquet
        ruta_parquet = carpeta_salida / "control_practicantes_silver.parquet"
        df.write_parquet(ruta_parquet, compression="snappy")
        print(f"   ✓ Parquet guardado: {ruta_parquet.name}")
        
        # Guardar Excel
        ruta_excel = carpeta_salida / "control_practicantes_silver.xlsx"
        df.write_excel(ruta_excel)
        print(f"   ✓ Excel guardado: {ruta_excel.name}")
        
        return {
            'success': True,
            'parquet': ruta_parquet,
            'excel': ruta_excel,
            'registros': registros_procesados
        }
        
    except Exception as e:
        print(f"   ✗ Error: {e}")
        raise


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n✗ Error fatal: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)