"""
Script de transformación Bronze → Silver
Reporte: Movimiento de Renovaciones
Fecha: 2024
Arquitectura: Bronze-Silver-Gold

Extrae datos desde fila 3 (headers) y procesa registros desde fila 4 en adelante.
"""

import polars as pl
import openpyxl
from pathlib import Path
from datetime import datetime
import sys

def extraer_datos_bronze(archivo_excel: Path) -> pl.DataFrame:
    """
    Extrae datos del archivo Excel bronze usando openpyxl.
    Headers en fila 3, datos desde fila 4.
    Convierte todo a String inicialmente para evitar conflictos de tipos.
    """
    print(f"\n📂 Leyendo archivo: {archivo_excel.name}")
    
    wb = openpyxl.load_workbook(archivo_excel, data_only=True)
    sheet = wb.active
    
    # Extraer encabezados de fila 3
    headers = []
    for cell in sheet[3]:
        if cell.value:
            headers.append(str(cell.value).strip())
        else:
            headers.append(f"COLUMNA_{len(headers)+1}")
    
    print(f"✓ Encabezados extraídos: {len(headers)} columnas")
    
    # Extraer datos desde fila 4 - CONVERTIR TODO A STRING
    datos = []
    filas_procesadas = 0
    
    for row in sheet.iter_rows(min_row=4, values_only=True):
        # Verificar si la fila tiene algún dato
        if any(cell is not None for cell in row):
            row_data = []
            for cell in row[:len(headers)]:
                if cell is not None:
                    # Convertir fechas a string en formato ISO
                    if isinstance(cell, datetime):
                        row_data.append(cell.strftime("%Y-%m-%d"))
                    else:
                        row_data.append(str(cell).strip())
                else:
                    row_data.append(None)
            
            # Rellenar con None si faltan columnas
            while len(row_data) < len(headers):
                row_data.append(None)
            datos.append(row_data)
            filas_procesadas += 1
    
    wb.close()
    
    print(f"✓ Filas procesadas: {filas_procesadas}")
    
    # Crear DataFrame con schema de solo Strings
    schema_string = {col: pl.String for col in headers}
    df = pl.DataFrame(datos, schema=schema_string, orient="row")
    
    print(f"✓ DataFrame creado: {df.shape}")
    
    return df


def limpiar_y_transformar(df: pl.DataFrame) -> pl.DataFrame:
    """
    Limpia y transforma los datos para la capa silver.
    Convierte tipos desde String a los tipos apropiados.
    """
    print("\n🔧 Iniciando limpieza y transformación...")
    
    registros_iniciales = len(df)
    print(f"  ├─ Registros iniciales: {registros_iniciales}")
    
    # 1. Eliminar filas completamente vacías
    df_limpio = df.filter(
        pl.any_horizontal([pl.col(c).is_not_null() for c in df.columns])
    )
    
    # 2. Limpiar espacios en todas las columnas (son String)
    for col in df_limpio.columns:
        df_limpio = df_limpio.with_columns(
            pl.col(col).str.strip_chars().alias(col)
        )
    
    # 3. Convertir strings vacíos a null
    for col in df_limpio.columns:
        df_limpio = df_limpio.with_columns(
            pl.when(pl.col(col) == "").then(None).otherwise(pl.col(col)).alias(col)
        )
    
    print(f"  ├─ Limpieza de espacios completada")
    
    # 4. Convertir tipos de datos específicos
    print(f"  ├─ Convirtiendo tipos de datos...")
    
    # Columnas numéricas
    transformaciones = []
    if "AÑO" in df_limpio.columns:
        transformaciones.append(
            pl.col("AÑO").cast(pl.Int32, strict=False).alias("AÑO")
        )
    
    if "MES" in df_limpio.columns:
        transformaciones.append(
            pl.col("MES").cast(pl.Int32, strict=False).alias("MES")
        )
    
    # FECHA DE NAC. - Manejar múltiples formatos
    # El archivo bronze puede tener fechas en dos formatos:
    # 1. YYYY-MM-DD (fechas datetime convertidas a string)
    # 2. dd/mm/yyyy (celdas de texto con tabs/espacios)
    if "FECHA DE NAC." in df_limpio.columns:
        df_limpio = df_limpio.with_columns([
            pl.when(pl.col("FECHA DE NAC.").str.to_date("%Y-%m-%d", strict=False).is_not_null())
              .then(pl.col("FECHA DE NAC.").str.to_date("%Y-%m-%d", strict=False))
              .otherwise(pl.col("FECHA DE NAC.").str.to_date("%d/%m/%Y", strict=False))
              .alias("FECHA DE NAC.")
        ])
        print(f"  ├─ FECHA DE NAC. convertida (formatos YYYY-MM-DD y dd/mm/yyyy)")
    
    # Otras fechas (solo formato ISO)
    for col_fecha in ["FECH_INGR.", "TERMINO DE CONTRATOS"]:
        if col_fecha in df_limpio.columns:
            transformaciones.append(
                pl.col(col_fecha).str.to_date("%Y-%m-%d", strict=False).alias(col_fecha)
            )
    
    # Aplicar transformaciones numéricas
    if transformaciones:
        df_limpio = df_limpio.with_columns(transformaciones)
        print(f"  ├─ Tipos de datos numéricos convertidos")
    
    # 5. Eliminar registros sin identificadores clave (DNI o CODIGO SAP2)
    registros_antes_filtro = len(df_limpio)
    
    if "DNI" in df_limpio.columns and "CODIGO SAP2" in df_limpio.columns:
        df_limpio = df_limpio.filter(
            (pl.col("DNI").is_not_null()) | (pl.col("CODIGO SAP2").is_not_null())
        )
    
    registros_eliminados_filtro = registros_antes_filtro - len(df_limpio)
    
    registros_finales = len(df_limpio)
    registros_eliminados = registros_iniciales - registros_finales
    
    print(f"  ├─ Registros eliminados sin DNI/SAP: {registros_eliminados_filtro}")
    print(f"  ├─ Registros después de limpieza: {registros_finales}")
    print(f"  └─ Total registros eliminados: {registros_eliminados}")
    
    return df_limpio


def generar_reportes_calidad(df_original: pl.DataFrame, df_limpio: pl.DataFrame) -> dict:
    """
    Genera métricas de calidad de datos.
    """
    print("\n📊 Generando reporte de calidad...")
    
    metricas = {
        "registros_originales": len(df_original),
        "registros_finales": len(df_limpio),
        "registros_eliminados": len(df_original) - len(df_limpio),
        "tasa_validacion": round(len(df_limpio) / len(df_original) * 100, 2) if len(df_original) > 0 else 0,
        "columnas_totales": len(df_limpio.columns)
    }
    
    # Analizar valores nulos en columnas clave
    columnas_clave = ["DNI", "CODIGO SAP2", "NOMBRE", "AP. PATERNO", "CARGO", "FECHA DE NAC.", "CORREO "]
    nulos_importantes = {}
    
    print(f"  ├─ Registros procesados: {metricas['registros_finales']}/{metricas['registros_originales']}")
    print(f"  ├─ Tasa de validación: {metricas['tasa_validacion']}%")
    print(f"  └─ Columnas: {metricas['columnas_totales']}")
    
    print(f"\n  📋 Análisis de columnas clave:")
    for col in columnas_clave:
        if col in df_limpio.columns:
            nulos = df_limpio[col].null_count()
            porcentaje = round(nulos / len(df_limpio) * 100, 2)
            nulos_importantes[col] = {"cantidad": nulos, "porcentaje": porcentaje}
            
            status = "✓" if nulos == 0 else "⚠️"
            print(f"      {status} {col}: {nulos} nulos ({porcentaje}%)")
    
    metricas["columnas_con_nulos"] = nulos_importantes
    
    # Distribución por año/mes
    if "AÑO" in df_limpio.columns and "MES" in df_limpio.columns:
        print(f"\n  📈 Distribución por periodo:")
        dist = df_limpio.group_by(["AÑO", "MES"]).agg(pl.len().alias("Registros")).sort(["AÑO", "MES"])
        for row in dist.head(5).iter_rows():
            print(f"      • {row[0]}-{row[1]:02d}: {row[2]} registros")
        
        if len(dist) > 5:
            print(f"      • ... y {len(dist) - 5} periodos más")
    
    return metricas


def guardar_silver(df: pl.DataFrame, directorio_salida: Path, nombre_base: str):
    """
    Guarda el DataFrame en formato Parquet y Excel en la capa silver.
    """
    print(f"\n💾 Guardando archivos en capa silver...")
    
    directorio_salida.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime("%d.%m.%Y_%H.%M.%S")
    
    # Guardar Parquet
    archivo_parquet = directorio_salida / f"{nombre_base}_silver_{timestamp}.parquet"
    df.write_parquet(archivo_parquet, compression="snappy")
    print(f"  ✓ Parquet: {archivo_parquet.name}")
    
    # Guardar Excel
    archivo_excel = directorio_salida / f"{nombre_base}_silver_{timestamp}.xlsx"
    df.write_excel(archivo_excel)
    print(f"  ✓ Excel: {archivo_excel.name}")
    
    return archivo_parquet, archivo_excel


def main():
    """
    Función principal del script.
    """
    print("="*80)
    print("TRANSFORMACIÓN BRONZE → SILVER")
    print("Reporte: Movimiento de Renovaciones")
    print("="*80)
    
    # Configuración de rutas
    try:
        # Solicitar archivo bronze
        import tkinter as tk
        from tkinter import filedialog
        
        root = tk.Tk()
        root.withdraw()
        
        archivo_bronze = filedialog.askopenfilename(
            title="Seleccionar archivo Bronze (Movimiento de Renovaciones)",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not archivo_bronze:
            print("❌ No se seleccionó ningún archivo")
            return
        
        archivo_bronze = Path(archivo_bronze)
        
        # Solicitar directorio de salida
        directorio_salida = filedialog.askdirectory(
            title="Seleccionar directorio para archivos Silver"
        )
        
        if not directorio_salida:
            print("❌ No se seleccionó directorio de salida")
            return
        
        directorio_salida = Path(directorio_salida)
        
    except Exception as e:
        print(f"❌ Error en selección de archivos: {e}")
        return
    
    try:
        # 1. Extraer datos bronze
        df_bronze = extraer_datos_bronze(archivo_bronze)
        
        # 2. Limpiar y transformar
        df_silver = limpiar_y_transformar(df_bronze)
        
        # 3. Generar reporte de calidad
        metricas = generar_reportes_calidad(df_bronze, df_silver)
        
        # 4. Guardar en capa silver
        archivo_parquet, archivo_excel = guardar_silver(
            df_silver, 
            directorio_salida, 
            "movimiento_renovaciones"
        )
        
        # Resumen final
        print("\n" + "="*80)
        print("✅ TRANSFORMACIÓN COMPLETADA EXITOSAMENTE")
        print("="*80)
        print(f"📁 Archivos generados:")
        print(f"   • {archivo_parquet.name}")
        print(f"   • {archivo_excel.name}")
        print(f"\n📊 Estadísticas:")
        print(f"   • Registros procesados: {metricas['registros_finales']}")
        print(f"   • Tasa de validación: {metricas['tasa_validacion']}%")
        print(f"   • Columnas: {metricas['columnas_totales']}")
        print("="*80)
        
    except Exception as e:
        print(f"\n❌ ERROR durante la transformación: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()