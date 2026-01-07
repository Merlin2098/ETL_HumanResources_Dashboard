"""
API Wrapper para Step 2: Silver → Gold (Transformación y Validación)
Expone función limpia para integración con CLI
"""

import polars as pl
import json
from pathlib import Path
from datetime import datetime
from typing import Tuple, List, Dict, Any
import shutil


def aplicar_transformaciones_gold(df: pl.DataFrame, schema: Dict[str, Any], logger=None) -> pl.DataFrame:
    """
    Aplica transformaciones según el esquema gold
    
    Args:
        df: DataFrame Silver de entrada
        schema: Esquema JSON con definiciones de columnas
        logger: Logger opcional
        
    Returns:
        pl.DataFrame: DataFrame transformado
    """
    if logger:
        logger.info("Seleccionando columnas según esquema...")
    
    # Seleccionar solo las columnas que existen en el schema
    columnas_schema = list(schema["schema"].keys())
    columnas_disponibles = [col for col in columnas_schema if col in df.columns]
    columnas_faltantes = [col for col in columnas_schema if col not in df.columns]
    
    if logger:
        logger.info(f"Columnas encontradas: {len(columnas_disponibles)}/{len(columnas_schema)}")
        if columnas_faltantes:
            logger.warning(f"Columnas faltantes del schema: {', '.join(columnas_faltantes[:5])}")
    
    # Seleccionar columnas
    df = df.select(columnas_disponibles)
    
    # Aplicar tipos de datos según schema
    if logger:
        logger.info("Aplicando tipos de datos...")
    
    for col_name in columnas_disponibles:
        col_spec = schema["schema"][col_name]
        col_type = col_spec["type"]
        
        try:
            if col_type == "string":
                df = df.with_columns(pl.col(col_name).cast(pl.Utf8))
            elif col_type == "integer":
                df = df.with_columns(pl.col(col_name).cast(pl.Int64))
            elif col_type == "float":
                df = df.with_columns(pl.col(col_name).cast(pl.Float64))
            elif col_type == "date":
                df = df.with_columns(
                    pl.col(col_name).str.strptime(pl.Datetime, "%Y-%m-%d %H:%M:%S%.f", strict=False).cast(pl.Date)
                )
        except Exception as e:
            if logger:
                logger.warning(f"Error al convertir columna {col_name}: {e}")
    
    # Aplicar string trimming
    if logger:
        logger.info("Limpiando espacios en strings...")
    
    for col in df.columns:
        if df[col].dtype == pl.Utf8:
            df = df.with_columns(pl.col(col).str.strip_chars())
    
    return df


def agregar_nombre_mes(df: pl.DataFrame, logger=None) -> pl.DataFrame:
    """
    Agrega columna con nombre del mes en español
    
    Args:
        df: DataFrame con columna MES
        logger: Logger opcional
        
    Returns:
        pl.DataFrame: DataFrame con columna NOMBRE_MES
    """
    if logger:
        logger.info("Agregando nombre del mes en español...")
    
    df = df.with_columns(
        pl.when(pl.col("MES") == 1).then(pl.lit("Enero"))
        .when(pl.col("MES") == 2).then(pl.lit("Febrero"))
        .when(pl.col("MES") == 3).then(pl.lit("Marzo"))
        .when(pl.col("MES") == 4).then(pl.lit("Abril"))
        .when(pl.col("MES") == 5).then(pl.lit("Mayo"))
        .when(pl.col("MES") == 6).then(pl.lit("Junio"))
        .when(pl.col("MES") == 7).then(pl.lit("Julio"))
        .when(pl.col("MES") == 8).then(pl.lit("Agosto"))
        .when(pl.col("MES") == 9).then(pl.lit("Septiembre"))
        .when(pl.col("MES") == 10).then(pl.lit("Octubre"))
        .when(pl.col("MES") == 11).then(pl.lit("Noviembre"))
        .when(pl.col("MES") == 12).then(pl.lit("Diciembre"))
        .otherwise(pl.lit(None))
        .alias("NOMBRE_MES")
    )
    
    # Reordenar columnas para colocar NOMBRE_MES después de MES
    columnas = df.columns
    indice_mes = columnas.index("MES")
    
    nuevo_orden = (
        columnas[:indice_mes + 1] +
        ["NOMBRE_MES"] +
        [col for col in columnas[indice_mes + 1:] if col != "NOMBRE_MES"]
    )
    
    df = df.select(nuevo_orden)
    
    if logger:
        logger.info("Columna 'NOMBRE_MES' agregada correctamente")
    
    return df


def validar_constraints(df: pl.DataFrame, schema: Dict[str, Any], logger=None) -> Tuple[List[str], List[str]]:
    """
    Valida constraints del schema
    
    Args:
        df: DataFrame a validar
        schema: Esquema con constraints
        logger: Logger opcional
        
    Returns:
        tuple: (errores, warnings)
    """
    errores = []
    warnings = []
    
    if logger:
        logger.info("Validando constraints del esquema...")
    
    # Validar primary key
    pk_cols = schema["constraints"]["primary_key"]
    pk_cols_existentes = [col for col in pk_cols if col in df.columns]
    
    if len(pk_cols_existentes) != len(pk_cols):
        faltantes = set(pk_cols) - set(pk_cols_existentes)
        warnings.append(f"Columnas de primary key faltantes: {faltantes}")
        pk_cols = pk_cols_existentes
    
    if pk_cols:
        duplicados = df.group_by(pk_cols).agg(pl.len().alias("count")).filter(pl.col("count") > 1)
        
        if len(duplicados) > 0:
            errores.append(f"Se encontraron {len(duplicados)} registros duplicados en primary key: {pk_cols}")
            if logger:
                logger.error(f"Primary key con duplicados: {len(duplicados)} registros")
        else:
            if logger:
                logger.info(f"Primary key sin duplicados: {pk_cols}")
    
    # Validar nulls en columnas no nullable
    nulls_encontrados = False
    for col_name, col_spec in schema["schema"].items():
        if col_name not in df.columns:
            continue
            
        if not col_spec.get("nullable", True):
            nulls = df.filter(pl.col(col_name).is_null()).height
            if nulls > 0:
                errores.append(f"Columna '{col_name}' tiene {nulls} valores nulos (no permitidos)")
                nulls_encontrados = True
    
    if not nulls_encontrados and logger:
        logger.info("Validación de nulls correcta")
    
    # Validar valores permitidos
    for col_name, col_spec in schema["schema"].items():
        if col_name not in df.columns:
            continue
            
        if "allowed_values" in col_spec:
            valores_invalidos = df.filter(
                ~pl.col(col_name).is_in(col_spec["allowed_values"]) & 
                pl.col(col_name).is_not_null()
            )
            if len(valores_invalidos) > 0:
                valores_unicos = valores_invalidos[col_name].unique().to_list()[:5]
                warnings.append(f"Columna '{col_name}' tiene {len(valores_invalidos)} valores fuera del rango: {valores_unicos}")
    
    # Validar rangos numéricos
    for col_name, col_spec in schema["schema"].items():
        if col_name not in df.columns:
            continue
            
        if "min_value" in col_spec:
            fuera_rango = df.filter(
                (pl.col(col_name) < col_spec["min_value"]) & 
                pl.col(col_name).is_not_null()
            ).height
            if fuera_rango > 0:
                warnings.append(f"Columna '{col_name}' tiene {fuera_rango} valores menores al mínimo ({col_spec['min_value']})")
    
    if logger:
        if errores:
            logger.log_validation_result(False, f"{len(errores)} errores críticos encontrados")
        elif warnings:
            logger.log_validation_result(True, f"{len(warnings)} advertencias encontradas")
        else:
            logger.log_validation_result(True, "Todas las validaciones pasaron")
    
    return errores, warnings


def gestionar_versionamiento_gold(carpeta_base: Path, logger=None) -> Tuple[Path, Path]:
    """
    Gestiona carpetas actual/ e historico/ para versionamiento
    
    Args:
        carpeta_base: Carpeta base del proyecto
        logger: Logger opcional
        
    Returns:
        tuple: (carpeta_actual, carpeta_historico)
    """
    carpeta_gold = carpeta_base / "gold"
    carpeta_actual = carpeta_gold / "actual"
    carpeta_historico = carpeta_gold / "historico"
    
    # Crear carpetas si no existen
    carpeta_actual.mkdir(parents=True, exist_ok=True)
    carpeta_historico.mkdir(parents=True, exist_ok=True)
    
    # Mover archivos existentes a histórico si existen
    nombre_parquet = "Planilla Metso BI_Gold.parquet"
    nombre_excel = "Planilla Metso BI_Gold.xlsx"
    
    archivo_actual_parquet = carpeta_actual / nombre_parquet
    archivo_actual_excel = carpeta_actual / nombre_excel
    
    if archivo_actual_parquet.exists() or archivo_actual_excel.exists():
        if logger:
            logger.info("Archivando versión anterior a histórico...")
        
        timestamp = datetime.now().strftime("%d.%m.%Y_%H.%M.%S")
        
        if archivo_actual_parquet.exists():
            nombre_historico_parquet = f"Planilla Metso BI_Gold_{timestamp}.parquet"
            destino_parquet = carpeta_historico / nombre_historico_parquet
            shutil.move(str(archivo_actual_parquet), str(destino_parquet))
            if logger:
                logger.info(f"Parquet archivado: {nombre_historico_parquet}")
        
        if archivo_actual_excel.exists():
            nombre_historico_excel = f"Planilla Metso BI_Gold_{timestamp}.xlsx"
            destino_excel = carpeta_historico / nombre_historico_excel
            shutil.move(str(archivo_actual_excel), str(destino_excel))
            if logger:
                logger.info(f"Excel archivado: {nombre_historico_excel}")
    
    return carpeta_actual, carpeta_historico


def generar_excel_visualizacion(df: pl.DataFrame, ruta_salida: Path, logger=None):
    """
    Genera Excel con formato para visualización
    
    Args:
        df: DataFrame a exportar
        ruta_salida: Ruta del archivo Excel de salida
        logger: Logger opcional
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    if logger:
        logger.info("Generando Excel con formato...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Planilla Gold"
    
    # Convertir a pandas para usar dataframe_to_rows
    df_pandas = df.to_pandas()
    
    # Escribir datos
    for r_idx, row in enumerate(dataframe_to_rows(df_pandas, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Formato de encabezado
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
    
    # Ajustar anchos de columna
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Congelar primera fila
    ws.freeze_panes = "A2"
    
    # Guardar
    wb.save(ruta_salida)
    
    if logger:
        logger.info(f"Excel formateado guardado: {ruta_salida.name}")


def exportar_silver_to_gold(
    parquet_silver: Path,
    schema_json: Path,
    carpeta_base: Path,
    skip_validation: bool = False,
    export_excel: bool = True,
    logger=None
) -> Tuple[pl.DataFrame, Path, Path]:
    """
    Transforma datos Silver a Gold con validación de esquema
    
    Args:
        parquet_silver: Ruta del archivo Parquet Silver
        schema_json: Ruta del esquema JSON Gold
        carpeta_base: Carpeta base del proyecto (se creará gold/actual y gold/historico)
        skip_validation: Si True, omite validaciones de constraints
        export_excel: Si True, genera también archivo Excel formateado
        logger: Logger opcional
        
    Returns:
        tuple: (df_gold, ruta_parquet_gold, ruta_excel_gold)
        
    Raises:
        Exception: Si hay errores críticos de validación
    """
    if logger:
        logger.log_step_start(
            "Transformación Silver → Gold",
            f"Procesando {parquet_silver.name}"
        )
    
    # Cargar datos Silver
    if logger:
        logger.info("Cargando datos Silver...")
    
    df = pl.read_parquet(parquet_silver)
    
    if logger:
        logger.log_dataframe_info("silver_input", rows=len(df), cols=len(df.columns))
    
    # Cargar schema
    if logger:
        logger.info(f"Cargando esquema: {schema_json.name}")
    
    with open(schema_json, 'r', encoding='utf-8') as f:
        schema = json.load(f)
    
    if logger:
        logger.info(f"Schema versión {schema['metadata']['version']}")
        logger.info(f"Columnas esperadas: {len(schema['schema'])}")
    
    # Aplicar transformaciones
    df_gold = aplicar_transformaciones_gold(df, schema, logger)
    
    # Agregar columna NOMBRE_MES
    df_gold = agregar_nombre_mes(df_gold, logger)
    
    if logger:
        logger.log_dataframe_info("gold_transformed", rows=len(df_gold), cols=len(df_gold.columns))
    
    # Validar constraints
    if not skip_validation:
        errores, warnings = validar_constraints(df_gold, schema, logger)
        
        if errores:
            for error in errores:
                if logger:
                    logger.error(error)
            raise Exception(f"Validación falló con {len(errores)} errores críticos")
        
        if warnings and logger:
            for warning in warnings:
                logger.warning(warning)
    else:
        if logger:
            logger.info("Validación de constraints omitida (skip_validation=True)")
    
    # Gestionar versionamiento
    carpeta_actual, carpeta_historico = gestionar_versionamiento_gold(carpeta_base, logger)
    
    # Rutas de salida
    ruta_parquet_gold = carpeta_actual / "Planilla Metso BI_Gold.parquet"
    ruta_excel_gold = carpeta_actual / "Planilla Metso BI_Gold.xlsx"
    
    # Guardar Parquet
    if logger:
        logger.info("Guardando Parquet Gold...")
    
    df_gold.write_parquet(ruta_parquet_gold)
    
    if logger:
        logger.log_file_processing(ruta_parquet_gold, "Guardado")
    
    # Guardar Excel si se solicita
    if export_excel:
        try:
            generar_excel_visualizacion(df_gold, ruta_excel_gold, logger)
            if logger:
                logger.log_file_processing(ruta_excel_gold, "Guardado")
        except Exception as e:
            if logger:
                logger.error(f"Error al generar Excel: {e}")
            ruta_excel_gold = None
    else:
        ruta_excel_gold = None
    
    if logger:
        logger.log_step_end("Transformación Silver → Gold", success=True)
    
    return df_gold, ruta_parquet_gold, ruta_excel_gold