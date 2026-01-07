"""
API Wrapper para Step 1: Bronze → Silver (Consolidación de Planillas)
Expone función limpia para integración con CLI
"""

import polars as pl
from pathlib import Path
from typing import List, Tuple, Optional
import re
from openpyxl import load_workbook


def extraer_periodo(nombre_archivo: str) -> Optional[str]:
    """
    Extrae el periodo del nombre del archivo
    
    Patrón esperado: METSO_Planilla YYYY-MM Empleados.xlsx
    
    Args:
        nombre_archivo: Nombre del archivo (str)
        
    Returns:
        str: Periodo en formato YYYY-MM o None si no se encuentra
    """
    patron = r'(\d{4}-\d{2})'
    match = re.search(patron, nombre_archivo)
    
    if match:
        return match.group(1)
    else:
        return None


def leer_archivo_planilla(archivo_path: Path, periodo: str, logger=None) -> pl.DataFrame:
    """
    Lee un archivo de planilla y agrega la columna PERIODO
    
    Args:
        archivo_path: Path del archivo Excel
        periodo: Periodo a agregar (str)
        logger: Logger opcional para registrar eventos
        
    Returns:
        pl.DataFrame: DataFrame con los datos del archivo
    """
    try:
        # Leer con openpyxl para obtener encabezados correctos
        wb = load_workbook(archivo_path, data_only=True, read_only=True)
        
        # Leer específicamente la hoja "Planilla"
        if "Planilla" not in wb.sheetnames:
            raise Exception(f"La hoja 'Planilla' no existe en el archivo. Hojas disponibles: {', '.join(wb.sheetnames)}")
        
        ws = wb["Planilla"]
        
        # Leer encabezados desde fila 6
        encabezados = []
        for col in range(1, 200):
            cell = ws.cell(row=6, column=col)
            valor = cell.value
            
            if valor is not None:
                encabezado = str(valor).strip()
                encabezados.append(encabezado)
            else:
                if col > len(encabezados) + 3:
                    break
        
        # Leer datos desde fila 7 en adelante
        datos = []
        for row in ws.iter_rows(min_row=7, values_only=True):
            if any(cell is not None for cell in row):
                primera_celda = row[0] if len(row) > 0 else None
                
                if primera_celda is None:
                    continue
                if isinstance(primera_celda, str) and len(primera_celda) > 20:
                    continue
                    
                datos.append(list(row[:len(encabezados)]))
        
        wb.close()
        
        # Crear diccionario para DataFrame
        data_dict = {encabezados[i]: [fila[i] if i < len(fila) else None for fila in datos] 
                     for i in range(len(encabezados))}
        
        # Crear DataFrame con polars
        try:
            df = pl.DataFrame(data_dict, strict=False)
        except Exception:
            data_dict_str = {k: [str(v) if v is not None else None for v in valores] 
                           for k, valores in data_dict.items()}
            df = pl.DataFrame(data_dict_str, strict=False)
        
        # Agregar columna PERIODO al inicio
        df = df.with_columns(
            pl.lit(periodo).alias("PERIODO")
        )
        
        # Reorganizar para que PERIODO sea la primera columna
        columnas = ["PERIODO"] + [col for col in df.columns if col != "PERIODO"]
        df = df.select(columnas)
        
        if logger:
            logger.info(f"Archivo leído: {archivo_path.name} ({len(df)} registros)")
        
        return df
        
    except Exception as e:
        if logger:
            logger.error(f"Error al leer {archivo_path.name}: {str(e)}")
        raise Exception(f"Error al leer {archivo_path.name}: {str(e)}")


def consolidar_planillas_bronze_to_silver(
    archivos_bronze: List[Path],
    output_dir: Path,
    logger=None
) -> Tuple[pl.DataFrame, Path, Path]:
    """
    Consolida múltiples archivos Excel de planillas en capa Silver
    
    Args:
        archivos_bronze: Lista de Path de archivos Excel a consolidar
        output_dir: Directorio donde guardar los resultados (se creará carpeta silver/)
        logger: Logger opcional para registrar eventos
        
    Returns:
        tuple: (df_consolidado, ruta_parquet, ruta_excel)
        
    Raises:
        Exception: Si no se puede procesar ningún archivo
    """
    dataframes = []
    archivos_procesados = 0
    archivos_con_error = []
    
    if logger:
        logger.log_step_start(
            "Consolidación de Archivos",
            f"Procesando {len(archivos_bronze)} archivo(s)"
        )
    
    # Procesar cada archivo
    for idx, archivo in enumerate(archivos_bronze, 1):
        try:
            # Extraer periodo del nombre del archivo
            periodo = extraer_periodo(archivo.name)
            
            if periodo is None:
                msg = f"No se pudo extraer periodo de {archivo.name}"
                if logger:
                    logger.warning(msg)
                archivos_con_error.append((archivo.name, "No se pudo extraer periodo"))
                continue
            
            if logger:
                logger.info(f"[{idx}/{len(archivos_bronze)}] Procesando {archivo.name} (Periodo: {periodo})")
            
            # Leer archivo
            df = leer_archivo_planilla(archivo, periodo, logger)
            dataframes.append(df)
            archivos_procesados += 1
            
            if logger:
                logger.log_dataframe_info(f"archivo_{idx}", rows=len(df), cols=len(df.columns))
            
        except Exception as e:
            if logger:
                logger.error(f"Error procesando {archivo.name}: {str(e)}")
            archivos_con_error.append((archivo.name, str(e)))
    
    if not dataframes:
        raise Exception("No se pudo procesar ningún archivo correctamente")
    
    if logger:
        logger.info(f"Archivos procesados exitosamente: {archivos_procesados}/{len(archivos_bronze)}")
        if archivos_con_error:
            logger.warning(f"Archivos con error: {len(archivos_con_error)}")
    
    # Consolidar todos los DataFrames
    if logger:
        logger.info("Normalizando tipos de datos...")
    
    dataframes_normalizados = []
    for df in dataframes:
        # Convertir todas las columnas (excepto PERIODO) a string
        columnas_a_convertir = [col for col in df.columns if col != "PERIODO"]
        df_normalizado = df.with_columns([
            pl.col(col).cast(pl.Utf8, strict=False).alias(col)
            for col in columnas_a_convertir
        ])
        dataframes_normalizados.append(df_normalizado)
    
    if logger:
        logger.info("Concatenando archivos...")
    
    df_consolidado = pl.concat(dataframes_normalizados, how="diagonal")
    
    if logger:
        logger.log_dataframe_info("consolidado", rows=len(df_consolidado), cols=len(df_consolidado.columns))
    
    # Limpieza: Eliminar filas con DNI/CEX nulo
    if logger:
        logger.info("Limpiando datos (DNI/CEX nulos)...")
    
    registros_antes = len(df_consolidado)
    
    if "DNI/CEX" in df_consolidado.columns:
        df_consolidado = df_consolidado.filter(pl.col("DNI/CEX").is_not_null())
        registros_despues = len(df_consolidado)
        registros_eliminados = registros_antes - registros_despues
        
        if registros_eliminados > 0 and logger:
            logger.info(f"Eliminadas {registros_eliminados:,} filas con DNI/CEX nulo")
    else:
        if logger:
            logger.warning("Columna 'DNI/CEX' no encontrada para limpieza")
    
    # Agregar columnas MES y AÑO derivadas de PERIODO
    if logger:
        logger.info("Generando columnas MES y AÑO desde PERIODO...")
    
    df_consolidado = df_consolidado.with_columns([
        pl.col("PERIODO").str.slice(0, 4).alias("AÑO"),
        pl.col("PERIODO").str.slice(5, 2).cast(pl.Int32, strict=False).alias("MES")
    ])
    
    # Reorganizar columnas: PERIODO, AÑO, MES, resto
    columnas_ordenadas = ["PERIODO", "AÑO", "MES"] + [
        col for col in df_consolidado.columns if col not in ["PERIODO", "AÑO", "MES"]
    ]
    df_consolidado = df_consolidado.select(columnas_ordenadas)
    
    # Crear carpeta silver si no existe
    carpeta_silver = Path(output_dir) / "silver"
    carpeta_silver.mkdir(exist_ok=True, parents=True)
    
    # Nombres fijos sin timestamp
    nombre_parquet = "Planilla Metso Consolidado.parquet"
    nombre_excel = "Planilla Metso Consolidado.xlsx"
    
    ruta_parquet = carpeta_silver / nombre_parquet
    ruta_excel = carpeta_silver / nombre_excel
    
    if logger:
        logger.log_step_start(
            "Exportación a Silver",
            f"Guardando resultados en {carpeta_silver}"
        )
    
    # Guardar como Parquet
    df_consolidado.write_parquet(ruta_parquet)
    if logger:
        logger.log_file_processing(ruta_parquet, "Guardado")
    
    # Guardar como Excel
    df_consolidado.write_excel(ruta_excel)
    if logger:
        logger.log_file_processing(ruta_excel, "Guardado")
    
    if logger:
        logger.log_step_end("Consolidación Bronze → Silver", success=True)
    
    return df_consolidado, ruta_parquet, ruta_excel