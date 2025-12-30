"""
Script: procesar_relacion_ingresos.py
Descripción: Procesa archivos Excel de Relación de Ingresos
             - Hoja EMPLEADOS: 22 columnas, headers en fila 2 (empieza en columna B)
             - Hoja PRACTICANTES: 20 columnas, headers en fila 2 (empieza en columna A)
             
Limpieza aplicada:
    - PROYECTO: Reemplazar "0" por null
    - CODIGO SAP: Reemplazar "#N/D" o "Error" por null
    
Autor: Richi
Fecha: 30.12.2024
"""

import polars as pl
import openpyxl
from pathlib import Path
from datetime import datetime
import sys
from tkinter import Tk, filedialog

# ============================================================================
# CONFIGURACIÓN
# ============================================================================

CONFIGURACION_HOJAS = {
    "EMPLEADOS": {
        "fila_headers": 2,
        "columna_inicio": 2,  # Columna B (1-indexed)
        "columnas_esperadas": [
            "AÑO", "MES", "DNI", "N° DOCUM.", "NOMBRES COMPLETOS", "SEXO",
            "DEPARTAMENTO", "PROVINCIA", "DISTRITO", "CARGO", "ÁREA", 
            "JEFE DIRECTO", "LUGAR DE TRABAJO", "FECHA INICIO", "CC",
            "WC/BC", "Tiempo de contrato", "NIVEL II", "NIVEL III", 
            "NIVEL IV", "PROYECTO", "CODIGO SAP"
        ]
    },
    "PRACTICANTES": {
        "fila_headers": 2,
        "columna_inicio": 1,  # Columna A (1-indexed)
        "columnas_esperadas": [
            "AÑO", "MES", "Tipo Documento", "Numero Documento", "CODIGO SAP",
            "Nombres Completos", "Fecha Nacimiento", "Sexo", "Departamento",
            "Provincia", "Distrito", "Nacionalidad", "Cargo", "Área",
            "Jefe responsable de área", "Sede de Trabajo", "Fecha Inicio",
            "CC", "Teléfono", "Universidad de Procedencia"
        ]
    }
}

# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================

def obtener_timestamp() -> str:
    """Genera timestamp en formato dd.mm.yyyy_hh.mm.ss"""
    return datetime.now().strftime("%d.%m.%Y_%H.%M.%S")


def seleccionar_archivo_excel() -> Path | None:
    """Abre diálogo para seleccionar archivo Excel"""
    root = Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    archivo = filedialog.askopenfilename(
        title="Seleccionar archivo de Relación de Ingresos (Bronze)",
        filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
    )
    
    root.destroy()
    
    return Path(archivo) if archivo else None


def leer_hoja_excel(ruta_excel: Path, nombre_hoja: str, config: dict) -> pl.DataFrame:
    """
    Lee una hoja Excel usando openpyxl con configuración específica
    
    Args:
        ruta_excel: Ruta al archivo Excel
        nombre_hoja: Nombre de la hoja a leer
        config: Diccionario con fila_headers y columna_inicio
    
    Returns:
        DataFrame de Polars con los datos
    """
    print(f"\n📖 Leyendo hoja: {nombre_hoja}")
    print(f"   Archivo: {ruta_excel.name}")
    print(f"   Headers en fila: {config['fila_headers']}, columna inicio: {config['columna_inicio']}")
    
    try:
        # Cargar workbook
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)
        
        if nombre_hoja not in wb.sheetnames:
            raise ValueError(f"La hoja '{nombre_hoja}' no existe. Hojas disponibles: {wb.sheetnames}")
        
        ws = wb[nombre_hoja]
        fila_headers = config['fila_headers']
        col_inicio = config['columna_inicio']
        
        # Extraer encabezados desde la columna especificada
        headers = []
        col_idx = col_inicio
        while True:
            cell = ws.cell(row=fila_headers, column=col_idx)
            if cell.value is None:
                break
            headers.append(str(cell.value).strip())
            col_idx += 1
        
        print(f"   ✓ Encabezados encontrados: {len(headers)} columnas")
        
        # Validar columnas esperadas
        columnas_esperadas = config['columnas_esperadas']
        if len(headers) != len(columnas_esperadas):
            print(f"   ⚠️  Advertencia: Se esperaban {len(columnas_esperadas)} columnas, se encontraron {len(headers)}")
        
        # Extraer datos
        datos = []
        fila_datos_inicio = fila_headers + 1
        
        for row in ws.iter_rows(min_row=fila_datos_inicio, values_only=True):
            # Tomar datos desde la columna de inicio
            fila_datos = list(row[col_inicio-1:col_inicio-1+len(headers)])
            
            # Detener si encontramos fila completamente vacía
            if all(cell is None or str(cell).strip() == "" for cell in fila_datos):
                break
            
            datos.append(fila_datos)
        
        wb.close()
        
        print(f"   ✓ Filas de datos extraídas: {len(datos)}")
        
        # Crear DataFrame de Polars
        if not datos:
            print("   ⚠️  No se encontraron datos")
            return pl.DataFrame()
        
        # Convertir todos los valores a string para evitar conflictos de tipo
        # (especialmente útil para fechas con formatos mixtos y números como strings)
        datos_str = []
        for fila in datos:
            fila_str = [str(valor) if valor is not None else None for valor in fila]
            datos_str.append(fila_str)
        
        # strict=False permite tipos mixtos en columnas
        df = pl.DataFrame(
            {header: [fila[i] for fila in datos_str] for i, header in enumerate(headers)},
            strict=False
        )
        
        return df
        
    except Exception as e:
        print(f"   ❌ Error al leer Excel: {str(e)}")
        raise


def limpiar_datos(df: pl.DataFrame, nombre_hoja: str) -> pl.DataFrame:
    """
    Aplica reglas de limpieza específicas para Relación de Ingresos
    
    Reglas:
    - PROYECTO: Reemplazar "0" por null
    - CODIGO SAP: Reemplazar "#N/D" o "Error" por null
    """
    print(f"\n🧹 Limpiando datos de {nombre_hoja}...")
    
    df_limpio = df
    stats = {"proyecto": 0, "codigo_sap": 0, "filas_vacias": 0}
    
    # Limpieza de PROYECTO (solo en EMPLEADOS)
    if "PROYECTO" in df.columns:
        registros_antes = df_limpio.filter(
            (pl.col("PROYECTO").cast(pl.Utf8) == "0") | 
            (pl.col("PROYECTO").is_null())
        ).height
        
        df_limpio = df_limpio.with_columns(
            pl.when(
                (pl.col("PROYECTO").cast(pl.Utf8) == "0") |
                (pl.col("PROYECTO").is_null())
            )
            .then(None)
            .otherwise(pl.col("PROYECTO"))
            .alias("PROYECTO")
        )
        
        stats["proyecto"] = registros_antes
        print(f"   ✓ PROYECTO: {registros_antes} valores '0' → null")
    
    # Limpieza de CODIGO SAP (ambas hojas)
    if "CODIGO SAP" in df.columns:
        # Contar registros que serán limpiados
        registros_antes = df_limpio.filter(
            pl.col("CODIGO SAP").cast(pl.Utf8).is_in(["#N/D", "Error", "#N/A"]) |
            pl.col("CODIGO SAP").cast(pl.Utf8).str.contains("(?i)error") |
            pl.col("CODIGO SAP").is_null()
        ).height
        
        df_limpio = df_limpio.with_columns(
            pl.when(
                (pl.col("CODIGO SAP").cast(pl.Utf8).str.contains("#N/D|#N/A")) |
                (pl.col("CODIGO SAP").cast(pl.Utf8).str.to_lowercase().str.contains("error")) |
                (pl.col("CODIGO SAP").is_null())
            )
            .then(None)
            .otherwise(pl.col("CODIGO SAP"))
            .alias("CODIGO SAP")
        )
        
        stats["codigo_sap"] = registros_antes
        print(f"   ✓ CODIGO SAP: {registros_antes} valores '#N/D'/'Error' → null")
    
    # Limpieza general: eliminar filas completamente vacías
    registros_antes = df_limpio.height
    df_limpio = df_limpio.filter(
        ~pl.all_horizontal(pl.all().is_null())
    )
    stats["filas_vacias"] = registros_antes - df_limpio.height
    
    if stats["filas_vacias"] > 0:
        print(f"   ✓ Eliminadas {stats['filas_vacias']} filas completamente vacías")
    
    return df_limpio


def generar_reporte_calidad(df_original: pl.DataFrame, df_limpio: pl.DataFrame, nombre_hoja: str):
    """Genera reporte de calidad de datos por hoja"""
    print(f"\n📊 REPORTE DE CALIDAD: {nombre_hoja}")
    print("=" * 80)
    print(f"Registros originales:  {df_original.height:,}")
    print(f"Registros limpios:     {df_limpio.height:,}")
    print(f"Registros eliminados:  {df_original.height - df_limpio.height:,}")
    
    if df_original.height > 0:
        print(f"Tasa de retención:     {(df_limpio.height / df_original.height * 100):.2f}%")
    
    print("=" * 80)
    
    # Estadísticas de nulos por columna (solo columnas con nulos)
    print("\n📋 Valores nulos por columna:")
    tiene_nulos = False
    for col in df_limpio.columns:
        nulos = df_limpio[col].is_null().sum()
        if nulos > 0:
            pct = (nulos / df_limpio.height * 100) if df_limpio.height > 0 else 0
            print(f"   {col:35} : {nulos:5,} ({pct:5.2f}%)")
            tiene_nulos = True
    
    if not tiene_nulos:
        print("   ✓ No hay valores nulos")


# ============================================================================
# FUNCIÓN PRINCIPAL
# ============================================================================

def main():
    """Función principal de procesamiento"""
    print("=" * 80)
    print("🚀 PROCESADOR DE RELACIÓN DE INGRESOS - BRONZE → SILVER")
    print("=" * 80)
    
    # 1. Seleccionar archivo
    archivo_bronze = seleccionar_archivo_excel()
    
    if not archivo_bronze:
        print("❌ No se seleccionó ningún archivo. Proceso cancelado.")
        return
    
    timestamp = obtener_timestamp()
    resultados = {}
    
    # 2. Procesar cada hoja
    for nombre_hoja, config in CONFIGURACION_HOJAS.items():
        print(f"\n{'='*80}")
        print(f"📄 PROCESANDO HOJA: {nombre_hoja}")
        print('='*80)
        
        try:
            # 2.1 Leer datos
            df_original = leer_hoja_excel(archivo_bronze, nombre_hoja, config)
            
            if df_original.is_empty():
                print(f"   ⚠️  No se encontraron datos en {nombre_hoja}")
                continue
            
            # 2.2 Limpiar datos
            df_limpio = limpiar_datos(df_original, nombre_hoja)
            
            # 2.3 Generar reporte de calidad
            generar_reporte_calidad(df_original, df_limpio, nombre_hoja)
            
            # 2.4 Guardar resultados
            resultados[nombre_hoja] = {
                "df": df_limpio,
                "registros": df_limpio.height
            }
            
        except Exception as e:
            print(f"\n❌ Error procesando {nombre_hoja}: {str(e)}")
            continue
    
    # 3. Guardar archivos procesados
    if not resultados:
        print("\n❌ No se procesaron datos de ninguna hoja.")
        return
    
    print(f"\n{'='*80}")
    print("💾 GUARDANDO ARCHIVOS PROCESADOS")
    print('='*80)
    
    # Obtener la carpeta donde está el archivo de origen
    carpeta_origen = archivo_bronze.parent
    print(f"\nCarpeta de destino: {carpeta_origen}")
    
    for nombre_hoja, datos in resultados.items():
        # Crear subcarpeta para cada hoja en la misma ubicación del archivo
        carpeta_salida = carpeta_origen / nombre_hoja.lower()
        carpeta_salida.mkdir(parents=True, exist_ok=True)
        
        # Guardar Parquet
        ruta_parquet = carpeta_salida / f"{nombre_hoja.lower()}_silver_{timestamp}.parquet"
        datos["df"].write_parquet(ruta_parquet, compression="snappy")
        print(f"\n✅ {nombre_hoja}")
        print(f"   Parquet: {ruta_parquet}")
        
        # Guardar Excel
        ruta_excel = carpeta_salida / f"{nombre_hoja.lower()}_silver_{timestamp}.xlsx"
        datos["df"].write_excel(ruta_excel)
        print(f"   Excel:   {ruta_excel}")
        print(f"   Registros: {datos['registros']:,}")
    
    # 4. Resumen final
    print("\n" + "=" * 80)
    print("✨ PROCESO COMPLETADO EXITOSAMENTE")
    print("=" * 80)
    print(f"\nHojas procesadas: {len(resultados)}")
    print(f"Total de registros: {sum(d['registros'] for d in resultados.values()):,}")
    print(f"\nArchivos generados en: {carpeta_origen}")
    for nombre_hoja in resultados.keys():
        print(f"  → {nombre_hoja.lower()}/")
    print("=" * 80)


# ============================================================================
# PUNTO DE ENTRADA
# ============================================================================

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ Error fatal: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)